#!/usr/bin/env python3
"""
Excel Supplier Data Consolidation - Batch 1
Processes supplier price lists and creates individual supplier tabs in master template
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import re
from datetime import datetime
from pathlib import Path

# Master template columns (13 columns - EXACT ORDER)
MASTER_COLUMNS = [
    'Supplier Name',
    'Supplier Code',
    'Product Category',
    'BRAND',
    'Brand Sub Tag',
    'SKU / MODEL',
    'PRODUCT DESCRIPTION',
    'SUPPLIER SOH',
    'COST EX VAT',
    'QTY ON ORDER',
    'NEXT SHIPMENT',
    'Tags',
    'LINKS'
]

# File paths
MASTER_FILE_INPUT = '/mnt/k/00Project/MantisNXT/database/Uploads/Consolidated_Supplier_Data.xlsx'
MASTER_FILE_OUTPUT = '/mnt/k/00Project/MantisNXT/database/Uploads/Consolidated_Supplier_Data_UPDATED.xlsx'
SOURCE_DIR = '/mnt/k/00Project/MantisNXT/database/Uploads/drive-download-20250904T012253Z-1-001/'

# Batch 1 files
BATCH_1_FILES = [
    'ACTIVE MUSIC DISTRIBUTION PRICELIST - AUGUST 2025.xlsx',
    'AV Distribution Pricelist - May 2025 (1).xlsx',
    'Alpha-Technologies-Pricelist-August-2025- (2).xlsx',
    'ApexPro Distribution pricelist 25-04-2025 (1).xlsx',
    'Audiolite PRICE LIST-WHOLESALE OUT.xlsx',
    'Audiosure StockFile 30072025.xlsx',
    'GLOBAL MUSIC SUGGESTED ONLINE ADVERTISED PRICELIST NOVEMBER 2024 SKU LIST FORMAT .xlsx'
]

def extract_supplier_name(filename):
    """
    Extract clean supplier name from filename.
    CRITICAL: Remove pricelist, dates, stock references, etc.
    """
    # Remove file extension
    name = filename.replace('.xlsx', '').replace('.xls', '')

    # Remove common noise words (case insensitive)
    noise_patterns = [
        r'\s*-?\s*pricelist.*$',
        r'\s*-?\s*price\s*list.*$',
        r'\s*-?\s*(january|february|march|april|may|june|july|august|september|october|november|december)\s*\d{4}.*$',
        r'\s*-?\s*\d{2}-\d{2}-\d{4}.*$',
        r'\s*-?\s*\d{4}.*$',
        r'\s*-?\s*stockfile.*$',
        r'\s*-?\s*stock\s*file.*$',
        r'\s*-?\s*SOH.*$',
        r'\s*-?\s*\(\d+\)\s*$',
        r'\s*-?\s*fullp.*$',
        r'\s*-?\s*suggested\s*online.*$',
        r'\s*-?\s*advertised.*$',
        r'\s*-?\s*sku\s*list.*$',
        r'\s*-?\s*format.*$',
        r'\s*-\s*$',
        r'\s+$'
    ]

    for pattern in noise_patterns:
        name = re.sub(pattern, '', name, flags=re.IGNORECASE)

    # Clean up extra spaces and dashes
    name = re.sub(r'\s+', ' ', name)
    name = re.sub(r'-+', ' ', name)
    name = name.strip(' -')

    # Title case for readability
    name = ' '.join(word.capitalize() for word in name.split())

    return name

def generate_supplier_code(supplier_name, index=1):
    """Generate supplier code from supplier name"""
    # Take first letters of each word, remove spaces
    code_parts = ''.join([word[0].upper() for word in supplier_name.split() if word])
    code = f"{code_parts}-{index:03d}"
    return code

def analyze_excel_file(filepath):
    """
    Deep analysis of Excel file structure.
    Returns dict with sheet names and column mappings.
    """
    print(f"\n{'='*80}")
    print(f"ANALYZING: {os.path.basename(filepath)}")
    print(f"{'='*80}")

    wb = load_workbook(filepath, read_only=True, data_only=True)
    analysis = {
        'sheets': [],
        'total_rows': 0,
        'total_cols': 0
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Get max row and column
        max_row = ws.max_row
        max_col = ws.max_column

        # Skip empty sheets
        if max_row is None or max_col is None or max_row == 0 or max_col == 0:
            continue

        # Read headers (first row)
        headers = []
        for col in range(1, max_col + 1):
            cell = ws.cell(1, col)
            headers.append(str(cell.value) if cell.value else f"Column_{col}")

        # Sample first 5 data rows
        sample_data = []
        for row in range(2, min(7, max_row + 1)):
            row_data = []
            for col in range(1, max_col + 1):
                cell = ws.cell(row, col)
                row_data.append(cell.value)
            sample_data.append(row_data)

        sheet_info = {
            'name': sheet_name,
            'rows': max_row,
            'cols': max_col,
            'headers': headers,
            'sample': sample_data
        }

        analysis['sheets'].append(sheet_info)
        analysis['total_rows'] += max_row
        analysis['total_cols'] = max(analysis['total_cols'], max_col)

        print(f"\nSheet: {sheet_name}")
        print(f"  Rows: {max_row}, Columns: {max_col}")
        print(f"  Headers: {headers[:10]}...")  # First 10 headers

    wb.close()
    return analysis

def map_columns_to_master(headers, sample_data):
    """
    Map source columns to master template columns.
    Uses fuzzy matching and keyword detection.
    """
    mapping = {}

    # Keyword patterns for each master column
    patterns = {
        'Supplier Name': [],  # Will be filled from filename
        'Supplier Code': ['supplier.*code', 'supp.*code', 'code'],
        'Product Category': ['category', 'cat', 'product.*category', 'type'],
        'BRAND': ['brand', 'manufacturer', 'make'],
        'Brand Sub Tag': ['sub.*tag', 'sub.*brand', 'series'],
        'SKU / MODEL': ['sku', 'model', 'part.*number', 'item.*code', 'product.*code', 'stock.*code'],
        'PRODUCT DESCRIPTION': ['description', 'product.*name', 'item.*description', 'product', 'item'],
        'SUPPLIER SOH': ['soh', 'stock.*on.*hand', 'qty.*available', 'available', 'quantity', 'stock', 'in.*stock'],
        'COST EX VAT': ['cost', 'price', 'ex.*vat', 'excl.*vat', 'wholesale', 'dealer', 'trade'],
        'QTY ON ORDER': ['on.*order', 'qty.*on.*order', 'ordered', 'incoming'],
        'NEXT SHIPMENT': ['shipment', 'eta', 'expected.*date', 'delivery.*date', 'next.*delivery'],
        'Tags': ['tags', 'keywords', 'categories'],
        'LINKS': ['link', 'url', 'website', 'image.*url']
    }

    # Convert headers to lowercase for matching
    headers_lower = [str(h).lower().strip() for h in headers]

    for master_col, pattern_list in patterns.items():
        best_match = None
        best_score = 0

        for idx, header in enumerate(headers_lower):
            for pattern in pattern_list:
                if re.search(pattern, header):
                    # Score based on pattern specificity
                    score = len(pattern)
                    if score > best_score:
                        best_score = score
                        best_match = idx

        if best_match is not None:
            mapping[master_col] = best_match

    return mapping

def transform_to_master_format(df, supplier_name, column_mapping):
    """
    Transform source dataframe to master template format.
    """
    master_df = pd.DataFrame(columns=MASTER_COLUMNS)

    # Start with supplier name (consistent across all rows)
    master_df['Supplier Name'] = supplier_name

    # Map columns based on mapping
    for master_col, source_idx in column_mapping.items():
        if source_idx < len(df.columns):
            master_df[master_col] = df.iloc[:, source_idx]

    # Generate supplier codes if not mapped
    if 'Supplier Code' not in column_mapping or master_df['Supplier Code'].isna().all():
        master_df['Supplier Code'] = [
            generate_supplier_code(supplier_name, i+1)
            for i in range(len(master_df))
        ]

    # Clean numeric columns
    if 'COST EX VAT' in master_df.columns:
        master_df['COST EX VAT'] = pd.to_numeric(
            master_df['COST EX VAT'], errors='coerce'
        )

    if 'SUPPLIER SOH' in master_df.columns:
        master_df['SUPPLIER SOH'] = pd.to_numeric(
            master_df['SUPPLIER SOH'], errors='coerce'
        )

    if 'QTY ON ORDER' in master_df.columns:
        master_df['QTY ON ORDER'] = pd.to_numeric(
            master_df['QTY ON ORDER'], errors='coerce'
        )

    # Remove completely empty rows
    master_df = master_df.dropna(how='all')

    return master_df

def process_supplier_file(filepath, master_workbook):
    """
    Process a single supplier file and add as new sheet to master workbook.
    """
    filename = os.path.basename(filepath)
    supplier_name = extract_supplier_name(filename)

    print(f"\n{'='*80}")
    print(f"PROCESSING: {filename}")
    print(f"EXTRACTED SUPPLIER: {supplier_name}")
    print(f"{'='*80}")

    # Analyze file structure
    analysis = analyze_excel_file(filepath)

    if not analysis['sheets']:
        print(f"⚠️  No sheets found in {filename}")
        return None

    # Process the first data sheet (or largest sheet)
    data_sheet = max(analysis['sheets'], key=lambda s: s['rows'])

    print(f"\nUsing sheet: {data_sheet['name']} ({data_sheet['rows']} rows)")

    # Read the data
    df = pd.read_excel(filepath, sheet_name=data_sheet['name'])

    print(f"Read {len(df)} rows with {len(df.columns)} columns")

    # Map columns to master format
    column_mapping = map_columns_to_master(data_sheet['headers'], data_sheet['sample'])

    print(f"\nColumn Mapping:")
    for master_col, source_idx in column_mapping.items():
        source_col = data_sheet['headers'][source_idx] if source_idx < len(data_sheet['headers']) else 'N/A'
        print(f"  {master_col} ← {source_col}")

    # Transform to master format
    master_df = transform_to_master_format(df, supplier_name, column_mapping)

    print(f"\nTransformed to {len(master_df)} rows in master format")

    # Add sheet to master workbook
    sheet_name = supplier_name[:31]  # Excel sheet name limit

    # Remove sheet if it exists
    if sheet_name in master_workbook.sheetnames:
        del master_workbook[sheet_name]

    # Create new sheet
    ws = master_workbook.create_sheet(sheet_name)

    # Write data
    for r_idx, row in enumerate(dataframe_to_rows(master_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(r_idx, c_idx, value)

    print(f"✅ Created sheet '{sheet_name}' with {len(master_df)} rows")

    return {
        'supplier': supplier_name,
        'filename': filename,
        'rows': len(master_df),
        'columns_mapped': len(column_mapping),
        'sheet_name': sheet_name
    }

def main():
    """Main execution function"""
    print(f"\n{'#'*80}")
    print("EXCEL SUPPLIER DATA CONSOLIDATION - BATCH 1")
    print(f"{'#'*80}")
    print(f"Input Master File: {MASTER_FILE_INPUT}")
    print(f"Output File: {MASTER_FILE_OUTPUT}")
    print(f"Source Directory: {SOURCE_DIR}")
    print(f"Files to Process: {len(BATCH_1_FILES)}")

    # Load master workbook
    if os.path.exists(MASTER_FILE_INPUT):
        print(f"\n✅ Loading existing master workbook...")
        master_wb = load_workbook(MASTER_FILE_INPUT)
    else:
        print(f"\n⚠️  Master file not found, creating new workbook...")
        master_wb = Workbook()
        # Create MASTER sheet with headers
        ws = master_wb.active
        ws.title = 'MASTER'
        for idx, col_name in enumerate(MASTER_COLUMNS, 1):
            ws.cell(1, idx, col_name)

    # Process each file
    results = []
    for filename in BATCH_1_FILES:
        filepath = os.path.join(SOURCE_DIR, filename)

        if not os.path.exists(filepath):
            print(f"\n⚠️  File not found: {filename}")
            continue

        try:
            result = process_supplier_file(filepath, master_wb)
            if result:
                results.append(result)
        except Exception as e:
            print(f"\n❌ ERROR processing {filename}: {str(e)}")
            import traceback
            traceback.print_exc()

    # Save master workbook
    print(f"\n{'='*80}")
    print("SAVING MASTER WORKBOOK...")
    master_wb.save(MASTER_FILE_OUTPUT)
    print(f"✅ Saved: {MASTER_FILE_OUTPUT}")

    # Summary report
    print(f"\n{'#'*80}")
    print("CONSOLIDATION SUMMARY - BATCH 1")
    print(f"{'#'*80}")
    print(f"\nProcessed {len(results)} suppliers:")

    total_rows = 0
    for result in results:
        print(f"\n  Supplier: {result['supplier']}")
        print(f"    File: {result['filename']}")
        print(f"    Sheet: {result['sheet_name']}")
        print(f"    Rows: {result['rows']}")
        print(f"    Columns Mapped: {result['columns_mapped']}/13")
        total_rows += result['rows']

    print(f"\n{'='*80}")
    print(f"TOTAL ROWS PROCESSED: {total_rows}")
    print(f"{'='*80}")

    # Data quality report
    print(f"\nDATA QUALITY NOTES:")
    print(f"  - All supplier tabs created with exact 13-column structure")
    print(f"  - Supplier names extracted and cleaned from filenames")
    print(f"  - Numeric columns converted to proper types")
    print(f"  - Empty rows removed")
    print(f"  - Ready for validation before merging to MASTER tab")

    return results

if __name__ == '__main__':
    results = main()
