#!/usr/bin/env python3
"""
Deep Excel File Inspector
Manually examines problematic supplier files to understand their exact structure
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os

SOURCE_DIR = '/mnt/k/00Project/MantisNXT/database/Uploads/drive-download-20250904T012253Z-1-001/'

PROBLEM_FILES = [
    'AV Distribution Pricelist - May 2025 (1).xlsx',
    'Alpha-Technologies-Pricelist-August-2025- (2).xlsx',
    'Audiolite PRICE LIST-WHOLESALE OUT.xlsx',
    'Audiosure StockFile 30072025.xlsx'
]

def deep_inspect_file(filepath):
    """Detailed inspection of Excel file"""
    print(f"\n{'='*100}")
    print(f"DEEP INSPECTION: {os.path.basename(filepath)}")
    print(f"{'='*100}")

    wb = load_workbook(filepath, read_only=False, data_only=True)

    for sheet_idx, sheet_name in enumerate(wb.sheetnames[:3]):  # First 3 sheets
        print(f"\n{'─'*100}")
        print(f"Sheet {sheet_idx + 1}: {sheet_name}")
        print(f"{'─'*100}")

        ws = wb[sheet_name]

        # Get dimensions
        max_row = ws.max_row
        max_col = ws.max_column

        print(f"Dimensions: {max_row} rows x {max_col} columns")

        # Show first 10 rows with all columns
        print(f"\nFirst 10 rows (showing all data):\n")

        for row_idx in range(1, min(11, max_row + 1)):
            row_data = []
            for col_idx in range(1, min(max_col + 1, 10)):  # First 10 columns
                cell = ws.cell(row_idx, col_idx)
                value = str(cell.value) if cell.value is not None else ''
                # Truncate long values
                if len(value) > 30:
                    value = value[:27] + '...'
                row_data.append(value)

            print(f"Row {row_idx:2d}: {' | '.join(row_data)}")

        # Try to identify header row
        print(f"\n{'-'*50}")
        print("Header Detection:")
        print(f"{'-'*50}")

        for check_row in range(1, min(6, max_row + 1)):
            row_values = []
            for col in range(1, min(max_col + 1, 10)):
                cell = ws.cell(check_row, col)
                if cell.value:
                    row_values.append(str(cell.value))

            if row_values:
                print(f"Row {check_row}: {row_values}")

                # Check if it looks like a header
                text_count = sum(1 for v in row_values if not v.replace('.','').replace('-','').isdigit())
                if text_count >= len(row_values) * 0.6:  # At least 60% text
                    print(f"  → Likely header row (60%+ text fields)")

    wb.close()

def main():
    """Main execution"""
    print(f"\n{'#'*100}")
    print("DEEP EXCEL FILE INSPECTOR")
    print(f"{'#'*100}")

    for filename in PROBLEM_FILES:
        filepath = os.path.join(SOURCE_DIR, filename)

        if not os.path.exists(filepath):
            print(f"\n⚠️  File not found: {filename}")
            continue

        try:
            deep_inspect_file(filepath)
        except Exception as e:
            print(f"\n❌ ERROR inspecting {filename}: {str(e)}")
            import traceback
            traceback.print_exc()

    print(f"\n{'#'*100}")
    print("INSPECTION COMPLETE")
    print(f"{'#'*100}")

if __name__ == '__main__':
    main()
