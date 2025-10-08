#!/usr/bin/env python3
"""
Final Master Consolidation and Comprehensive Audit
Consolidates all validated supplier tabs into Master tab with full audit trail
"""

import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import json
from pathlib import Path
from collections import defaultdict

class MasterConsolidator:
    """Consolidates all supplier tabs into Master tab with comprehensive audit"""

    def __init__(self, file_path: str):
        self.file_path = file_path
        self.wb = None
        self.master_schema = {
            'required_columns': [
                'Product_Code', 'Category', 'Product_Name', 'Brand',
                'Supplier', 'Unit_Price', 'Quantity_Available',
                'Reorder_Level', 'Last_Updated'
            ],
            'optional_columns': [
                'Description', 'Unit_of_Measure', 'Lead_Time_Days',
                'Min_Order_Quantity', 'Discount_Available'
            ]
        }
        self.audit_results = {
            'consolidation_timestamp': datetime.now().isoformat(),
            'total_products': 0,
            'products_per_supplier': {},
            'data_quality': {},
            'issues': [],
            'warnings': [],
            'performance_metrics': {}
        }

    def load_workbook(self):
        """Load Excel workbook"""
        try:
            self.wb = openpyxl.load_workbook(self.file_path)
            print(f"✅ Loaded workbook: {self.file_path}")
            print(f"📋 Available sheets: {self.wb.sheetnames}")
            return True
        except Exception as e:
            print(f"❌ Error loading workbook: {e}")
            return False

    def identify_supplier_tabs(self) -> list:
        """Identify all validated supplier tabs"""
        supplier_tabs = []
        exclude_tabs = ['Master', 'MASTER', 'Audit_Log', 'Summary', 'Metadata',
                       'All_Products', 'Processing_Log']

        for sheet_name in self.wb.sheetnames:
            if sheet_name not in exclude_tabs:
                supplier_tabs.append(sheet_name)

        print(f"\n🔍 Identified {len(supplier_tabs)} supplier tabs:")
        for tab in supplier_tabs:
            print(f"   - {tab}")

        return supplier_tabs

    def read_supplier_data(self, sheet_name: str) -> pd.DataFrame:
        """Read data from supplier sheet"""
        try:
            df = pd.read_excel(self.file_path, sheet_name=sheet_name)

            # Add Supplier column if not present
            if 'Supplier' not in df.columns:
                df['Supplier'] = sheet_name

            return df
        except Exception as e:
            self.audit_results['issues'].append(f"Failed to read {sheet_name}: {e}")
            return pd.DataFrame()

    def validate_schema(self, df: pd.DataFrame, supplier: str) -> dict:
        """Validate data against master schema"""
        validation = {
            'supplier': supplier,
            'total_rows': len(df),
            'missing_required': [],
            'missing_optional': [],
            'completeness': {},
            'data_types': {},
            'issues': []
        }

        # Check required columns
        for col in self.master_schema['required_columns']:
            if col not in df.columns:
                validation['missing_required'].append(col)
            else:
                # Calculate completeness
                non_null = df[col].notna().sum()
                completeness = (non_null / len(df) * 100) if len(df) > 0 else 0
                validation['completeness'][col] = round(completeness, 2)

                # Check if completeness is too low for required fields
                if completeness < 80:
                    validation['issues'].append(
                        f"{col}: Only {completeness}% complete (below 80% threshold)"
                    )

        # Check optional columns
        for col in self.master_schema['optional_columns']:
            if col not in df.columns:
                validation['missing_optional'].append(col)
            else:
                non_null = df[col].notna().sum()
                completeness = (non_null / len(df) * 100) if len(df) > 0 else 0
                validation['completeness'][col] = round(completeness, 2)

        return validation

    def consolidate_data(self, supplier_tabs: list) -> pd.DataFrame:
        """Consolidate all supplier data into single DataFrame"""
        print("\n📦 Starting consolidation process...")

        all_data = []

        for supplier in supplier_tabs:
            print(f"\n   Processing: {supplier}")

            # Read supplier data
            df = self.read_supplier_data(supplier)

            if df.empty:
                print(f"      ⚠️  Empty or failed to read")
                continue

            # Validate schema
            validation = self.validate_schema(df, supplier)
            self.audit_results['data_quality'][supplier] = validation

            # Track product count
            self.audit_results['products_per_supplier'][supplier] = len(df)

            # Report issues
            if validation['missing_required']:
                print(f"      ⚠️  Missing required columns: {validation['missing_required']}")
                self.audit_results['warnings'].append(
                    f"{supplier}: Missing required columns {validation['missing_required']}"
                )

            if validation['issues']:
                print(f"      ⚠️  Data quality issues: {len(validation['issues'])}")
                for issue in validation['issues']:
                    print(f"         - {issue}")

            # Add to consolidation
            all_data.append(df)
            print(f"      ✅ Added {len(df)} products")

        # Combine all data
        if all_data:
            master_df = pd.concat(all_data, ignore_index=True)
            self.audit_results['total_products'] = len(master_df)
            print(f"\n✅ Consolidated {len(master_df)} total products from {len(all_data)} suppliers")
            return master_df
        else:
            print("\n❌ No data to consolidate")
            return pd.DataFrame()

    def remove_duplicates(self, df: pd.DataFrame) -> pd.DataFrame:
        """Remove duplicate products and track metrics"""
        initial_count = len(df)

        # Check for duplicates based on Product_Code and Supplier
        if 'Product_Code' in df.columns and 'Supplier' in df.columns:
            duplicates = df.duplicated(subset=['Product_Code', 'Supplier'], keep='first')
            duplicate_count = duplicates.sum()

            if duplicate_count > 0:
                print(f"\n🔍 Found {duplicate_count} duplicate products")
                df = df[~duplicates]
                self.audit_results['performance_metrics']['duplicates_removed'] = duplicate_count
            else:
                print("\n✅ No duplicates found")
                self.audit_results['performance_metrics']['duplicates_removed'] = 0

        final_count = len(df)
        self.audit_results['performance_metrics']['initial_count'] = initial_count
        self.audit_results['performance_metrics']['final_count'] = final_count

        return df

    def optimize_data_types(self, df: pd.DataFrame) -> pd.DataFrame:
        """Optimize data types for better performance"""
        print("\n🔧 Optimizing data types...")

        # Numeric columns
        numeric_cols = ['Unit_Price', 'Quantity_Available', 'Reorder_Level',
                       'Lead_Time_Days', 'Min_Order_Quantity']

        for col in numeric_cols:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')

        # Date columns
        if 'Last_Updated' in df.columns:
            df['Last_Updated'] = pd.to_datetime(df['Last_Updated'], errors='coerce')

        # String columns (category type for better memory)
        string_cols = ['Category', 'Brand', 'Supplier', 'Unit_of_Measure']
        for col in string_cols:
            if col in df.columns:
                df[col] = df[col].astype('category')

        print("   ✅ Data types optimized")
        return df

    def write_master_tab(self, df: pd.DataFrame):
        """Write consolidated data to Master tab with formatting"""
        print("\n📝 Writing to Master tab...")

        # Remove existing Master sheet if present (check both cases)
        master_sheet_name = None
        if 'MASTER' in self.wb.sheetnames:
            master_sheet_name = 'MASTER'
        elif 'Master' in self.wb.sheetnames:
            master_sheet_name = 'Master'

        if master_sheet_name:
            del self.wb[master_sheet_name]
            print(f"   🗑️  Removed existing {master_sheet_name} tab")

        # Create new Master sheet at first position
        self.wb.create_sheet('MASTER', 0)
        ws = self.wb['MASTER']

        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                # Header formatting
                if r_idx == 1:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = Border(
                        bottom=Side(style='thick', color="000000")
                    )

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze header row
        ws.freeze_panes = 'A2'

        print(f"   ✅ Written {len(df)} rows to Master tab")

    def calculate_coverage_analysis(self, df: pd.DataFrame) -> dict:
        """Calculate field coverage statistics"""
        coverage = {}

        for col in df.columns:
            non_null = df[col].notna().sum()
            total = len(df)
            percentage = (non_null / total * 100) if total > 0 else 0
            coverage[col] = {
                'filled': non_null,
                'total': total,
                'percentage': round(percentage, 2)
            }

        return coverage

    def create_summary_statistics(self, df: pd.DataFrame):
        """Create comprehensive summary statistics"""
        print("\n📊 Generating summary statistics...")

        summary = {
            'consolidation_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'total_products': len(df),
            'total_suppliers': df['Supplier'].nunique() if 'Supplier' in df.columns else 0,
            'products_per_supplier': self.audit_results['products_per_supplier'],
            'coverage_analysis': self.calculate_coverage_analysis(df),
            'category_breakdown': df['Category'].value_counts().to_dict() if 'Category' in df.columns else {},
            'brand_breakdown': df['Brand'].value_counts().to_dict() if 'Brand' in df.columns else {},
            'price_statistics': {
                'min': float(df['Unit_Price'].min()) if 'Unit_Price' in df.columns else 0,
                'max': float(df['Unit_Price'].max()) if 'Unit_Price' in df.columns else 0,
                'mean': float(df['Unit_Price'].mean()) if 'Unit_Price' in df.columns else 0,
                'median': float(df['Unit_Price'].median()) if 'Unit_Price' in df.columns else 0
            } if 'Unit_Price' in df.columns else {},
            'inventory_statistics': {
                'total_quantity': int(df['Quantity_Available'].sum()) if 'Quantity_Available' in df.columns else 0,
                'low_stock_items': int((df['Quantity_Available'] <= df['Reorder_Level']).sum()) if all(col in df.columns for col in ['Quantity_Available', 'Reorder_Level']) else 0
            }
        }

        self.audit_results['summary_statistics'] = summary

        # Print summary
        print(f"\n   📦 Total Products: {summary['total_products']:,}")
        print(f"   🏢 Total Suppliers: {summary['total_suppliers']}")
        print(f"   📊 Categories: {len(summary['category_breakdown'])}")
        print(f"   🏷️  Brands: {len(summary['brand_breakdown'])}")

        return summary

    def write_audit_sheet(self):
        """Create comprehensive audit sheet"""
        print("\n📋 Creating Audit sheet...")

        # Remove existing Audit sheet if present
        if 'Audit_Log' in self.wb.sheetnames:
            del self.wb['Audit_Log']

        self.wb.create_sheet('Audit_Log')
        ws = self.wb['Audit_Log']

        # Title
        ws['A1'] = 'MASTER CONSOLIDATION AUDIT REPORT'
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
        ws.merge_cells('A1:D1')

        row = 3

        # Consolidation Summary
        ws[f'A{row}'] = 'CONSOLIDATION SUMMARY'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        ws[f'A{row}'] = 'Timestamp:'
        ws[f'B{row}'] = self.audit_results['consolidation_timestamp']
        row += 1

        ws[f'A{row}'] = 'Total Products:'
        ws[f'B{row}'] = self.audit_results['total_products']
        row += 1

        ws[f'A{row}'] = 'Total Suppliers:'
        ws[f'B{row}'] = len(self.audit_results['products_per_supplier'])
        row += 2

        # Products per Supplier
        ws[f'A{row}'] = 'PRODUCTS PER SUPPLIER'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        ws[f'A{row}'] = 'Supplier'
        ws[f'B{row}'] = 'Product Count'
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        row += 1

        for supplier, count in sorted(self.audit_results['products_per_supplier'].items()):
            ws[f'A{row}'] = supplier
            ws[f'B{row}'] = count
            row += 1

        row += 1

        # Data Quality Summary
        ws[f'A{row}'] = 'DATA QUALITY SUMMARY'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        for supplier, quality in self.audit_results['data_quality'].items():
            ws[f'A{row}'] = supplier
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

            ws[f'A{row}'] = '  Total Rows:'
            ws[f'B{row}'] = quality['total_rows']
            row += 1

            if quality['missing_required']:
                ws[f'A{row}'] = '  Missing Required:'
                ws[f'B{row}'] = ', '.join(quality['missing_required'])
                ws[f'B{row}'].font = Font(color="FF0000")
                row += 1

            if quality['issues']:
                ws[f'A{row}'] = '  Issues:'
                row += 1
                for issue in quality['issues']:
                    ws[f'A{row}'] = f"    - {issue}"
                    ws[f'A{row}'].font = Font(color="FFA500")
                    row += 1

            row += 1

        # Performance Metrics
        if self.audit_results['performance_metrics']:
            ws[f'A{row}'] = 'PERFORMANCE METRICS'
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1

            for metric, value in self.audit_results['performance_metrics'].items():
                ws[f'A{row}'] = metric.replace('_', ' ').title() + ':'
                ws[f'B{row}'] = value
                row += 1

        # Auto-adjust columns
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 30

        print("   ✅ Audit sheet created")

    def write_summary_sheet(self):
        """Create executive summary sheet"""
        print("\n📊 Creating Summary sheet...")

        if 'Summary' in self.wb.sheetnames:
            del self.wb['Summary']

        self.wb.create_sheet('Summary')
        ws = self.wb['Summary']

        # Title
        ws['A1'] = 'EXECUTIVE SUMMARY - MASTER CONSOLIDATION'
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
        ws.merge_cells('A1:D1')

        row = 3

        summary = self.audit_results.get('summary_statistics', {})

        # Key Metrics
        ws[f'A{row}'] = 'KEY METRICS'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        metrics = [
            ('Consolidation Date:', summary.get('consolidation_date', 'N/A')),
            ('Total Products:', f"{summary.get('total_products', 0):,}"),
            ('Total Suppliers:', summary.get('total_suppliers', 0)),
            ('Total Categories:', len(summary.get('category_breakdown', {}))),
            ('Total Brands:', len(summary.get('brand_breakdown', {}))),
        ]

        for label, value in metrics:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

        row += 1

        # Price Statistics
        if 'price_statistics' in summary and summary['price_statistics']:
            ws[f'A{row}'] = 'PRICE STATISTICS'
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1

            price_stats = summary['price_statistics']
            stats = [
                ('Minimum Price:', f"${price_stats.get('min', 0):.2f}"),
                ('Maximum Price:', f"${price_stats.get('max', 0):.2f}"),
                ('Average Price:', f"${price_stats.get('mean', 0):.2f}"),
                ('Median Price:', f"${price_stats.get('median', 0):.2f}"),
            ]

            for label, value in stats:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                ws[f'A{row}'].font = Font(bold=True)
                row += 1

            row += 1

        # Inventory Statistics
        if 'inventory_statistics' in summary:
            ws[f'A{row}'] = 'INVENTORY STATISTICS'
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1

            inv_stats = summary['inventory_statistics']
            ws[f'A{row}'] = 'Total Quantity Available:'
            ws[f'B{row}'] = f"{inv_stats.get('total_quantity', 0):,}"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

            ws[f'A{row}'] = 'Low Stock Items:'
            ws[f'B{row}'] = inv_stats.get('low_stock_items', 0)
            ws[f'A{row}'].font = Font(bold=True)
            if inv_stats.get('low_stock_items', 0) > 0:
                ws[f'B{row}'].font = Font(color="FF0000", bold=True)
            row += 1

        # Auto-adjust columns
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

        print("   ✅ Summary sheet created")

    def save_workbook(self):
        """Save the workbook"""
        try:
            self.wb.save(self.file_path)
            print(f"\n✅ Workbook saved successfully: {self.file_path}")
            return True
        except Exception as e:
            print(f"\n❌ Error saving workbook: {e}")
            self.audit_results['issues'].append(f"Failed to save workbook: {e}")
            return False

    def save_audit_report(self):
        """Save audit results to JSON file"""
        report_path = Path(self.file_path).parent / 'consolidation_audit_report.json'

        try:
            with open(report_path, 'w', encoding='utf-8') as f:
                json.dump(self.audit_results, f, indent=2, default=str)

            print(f"\n✅ Audit report saved: {report_path}")
            return True
        except Exception as e:
            print(f"\n❌ Error saving audit report: {e}")
            return False

    def execute(self):
        """Execute full consolidation and audit process"""
        print("=" * 80)
        print("MASTER TAB CONSOLIDATION & COMPREHENSIVE AUDIT")
        print("=" * 80)

        # Load workbook
        if not self.load_workbook():
            return False

        # Identify supplier tabs
        supplier_tabs = self.identify_supplier_tabs()

        if not supplier_tabs:
            print("\n❌ No supplier tabs found to consolidate")
            return False

        # Consolidate data
        master_df = self.consolidate_data(supplier_tabs)

        if master_df.empty:
            print("\n❌ Consolidation failed - no data")
            return False

        # Remove duplicates
        master_df = self.remove_duplicates(master_df)

        # Optimize data types
        master_df = self.optimize_data_types(master_df)

        # Write to Master tab
        self.write_master_tab(master_df)

        # Create summary statistics
        self.create_summary_statistics(master_df)

        # Create audit sheet
        self.write_audit_sheet()

        # Create summary sheet
        self.write_summary_sheet()

        # Save workbook
        if not self.save_workbook():
            return False

        # Save audit report
        self.save_audit_report()

        # Final report
        print("\n" + "=" * 80)
        print("CONSOLIDATION COMPLETE")
        print("=" * 80)
        print(f"✅ Total Products: {self.audit_results['total_products']:,}")
        print(f"✅ Total Suppliers: {len(self.audit_results['products_per_supplier'])}")
        print(f"✅ Duplicates Removed: {self.audit_results['performance_metrics'].get('duplicates_removed', 0)}")

        if self.audit_results['warnings']:
            print(f"\n⚠️  Warnings: {len(self.audit_results['warnings'])}")
            for warning in self.audit_results['warnings'][:5]:
                print(f"   - {warning}")

        if self.audit_results['issues']:
            print(f"\n❌ Issues: {len(self.audit_results['issues'])}")
            for issue in self.audit_results['issues'][:5]:
                print(f"   - {issue}")

        print("\n✅ Consolidation and audit complete!")
        return True

def main():
    file_path = '/mnt/k/00Project/MantisNXT/database/Uploads/Consolidated_Supplier_Data.xlsx'

    consolidator = MasterConsolidator(file_path)
    success = consolidator.execute()

    if success:
        print("\n🎉 Master tab consolidation successful!")
        return 0
    else:
        print("\n❌ Master tab consolidation failed")
        return 1

if __name__ == '__main__':
    exit(main())
