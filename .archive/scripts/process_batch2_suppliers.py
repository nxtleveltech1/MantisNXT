#!/usr/bin/env python3
"""
Supplier Data Consolidation - Batch 2 Processor
Processes 7 supplier files and creates individual tabs in Consolidated_Supplier_Data.xlsx
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from pathlib import Path
import warnings
warnings.filterwarnings('ignore')

# Paths
SOURCE_DIR = Path('/mnt/k/00Project/MantisNXT/database/Uploads/drive-download-20250904T012253Z-1-001')
MASTER_FILE = Path('/mnt/k/00Project/MantisNXT/database/Uploads/Consolidated_Supplier_Data_Batch2.xlsx')

# Master template columns (exact order)
MASTER_COLUMNS = [
    'Supplier Name ',
    'Supplier Code',
    'Produt Category',
    'BRAND',
    'Brand Sub Tag',
    'SKU / MODEL ',
    'PRODUCT DESCRIPTION',
    'SUPPLIER SOH',
    'COST  EX VAT',
    'QTY ON ORDER',
    'NEXT SHIPMENT',
    'Tags',
    'LINKS'
]

# Batch 2 files with supplier name mappings
BATCH_2_FILES = {
    'MD External Stock 2025-08-25.xlsx': 'MD Distribution',
    'Music Power Pricelist (August 2025).xlsx': 'Music Power',
    'Planetworld SOH 20 June.xlsx': 'Planetworld',
    'Pro Audio platinum .xlsx': 'Pro Audio Platinum',
    'Rockit Price_List_25.08.2025.01.xlsx': 'Rockit',
    'Rolling Thunder July 2025 Pricelist .xlsx': 'Rolling Thunder',
    'Sennheiser 2025 (2).xlsx': 'Sennheiser'
}

class SupplierProcessor:
    def __init__(self, source_file, supplier_name):
        self.source_file = source_file
        self.supplier_name = supplier_name
        self.df_source = None
        self.df_master = None
        self.issues = []

    def analyze_file(self):
        """Deep analysis of source file structure"""
        print(f"\n{'='*80}")
        print(f"ANALYZING: {self.supplier_name}")
        print(f"File: {self.source_file.name}")
        print(f"{'='*80}")

        try:
            # Load Excel file
            xl = pd.ExcelFile(self.source_file)
            print(f"Sheets found: {xl.sheet_names}")

            # Try to find the main data sheet
            for sheet_name in xl.sheet_names:
                df = pd.read_excel(self.source_file, sheet_name=sheet_name)

                # Skip empty sheets
                if df.empty:
                    continue

                print(f"\nSheet: '{sheet_name}'")
                print(f"Shape: {df.shape}")
                print(f"Columns: {list(df.columns)[:10]}")  # First 10 columns

                # Look for data-rich sheet
                if df.shape[0] > 10:  # At least 10 rows
                    self.df_source = df
                    print(f"✓ Selected sheet '{sheet_name}' as main data source")
                    break

            if self.df_source is None:
                self.issues.append("No suitable data sheet found")
                return False

            return True

        except Exception as e:
            self.issues.append(f"File read error: {str(e)}")
            print(f"ERROR: {e}")
            return False

    def map_columns(self):
        """Intelligent column mapping to master format"""
        print(f"\nColumn Mapping for {self.supplier_name}:")
        print("-" * 80)

        # Initialize master dataframe
        self.df_master = pd.DataFrame(columns=MASTER_COLUMNS)

        # Get source columns (lowercase for matching)
        source_cols = {col.lower().strip(): col for col in self.df_source.columns if isinstance(col, str)}

        # Mapping logic
        mappings = {}

        # Supplier Name - always the supplier name
        mappings['Supplier Name '] = self.supplier_name

        # SKU / MODEL - look for variations
        for pattern in ['sku', 'model', 'code', 'part', 'item']:
            matches = [col for col in source_cols.keys() if pattern in col]
            if matches:
                mappings['SKU / MODEL '] = self.df_source[source_cols[matches[0]]]
                print(f"SKU / MODEL : {matches[0]}")
                break

        # PRODUCT DESCRIPTION
        for pattern in ['description', 'product', 'name', 'item']:
            matches = [col for col in source_cols.keys() if pattern in col and 'price' not in col]
            if matches:
                mappings['PRODUCT DESCRIPTION'] = self.df_source[source_cols[matches[0]]]
                print(f"PRODUCT DESCRIPTION: {matches[0]}")
                break

        # BRAND
        for pattern in ['brand', 'make', 'manufacturer']:
            matches = [col for col in source_cols.keys() if pattern in col]
            if matches:
                mappings['BRAND'] = self.df_source[source_cols[matches[0]]]
                print(f"BRAND: {matches[0]}")
                break

        # SUPPLIER SOH (Stock on Hand)
        for pattern in ['soh', 'stock', 'qty', 'quantity', 'available', 'on hand']:
            matches = [col for col in source_cols.keys() if pattern in col and 'price' not in col and 'cost' not in col]
            if matches:
                mappings['SUPPLIER SOH'] = self.df_source[source_cols[matches[0]]]
                print(f"SUPPLIER SOH: {matches[0]}")
                break

        # COST EX VAT
        for pattern in ['cost', 'price', 'ex vat', 'exvat', 'dealer', 'wholesale']:
            matches = [col for col in source_cols.keys() if pattern in col]
            if matches:
                mappings['COST  EX VAT'] = self.df_source[source_cols[matches[0]]]
                print(f"COST  EX VAT: {matches[0]}")
                break

        # Product Category
        for pattern in ['category', 'dept', 'department', 'type', 'class']:
            matches = [col for col in source_cols.keys() if pattern in col]
            if matches:
                mappings['Produt Category'] = self.df_source[source_cols[matches[0]]]
                print(f"Produt Category: {matches[0]}")
                break

        # QTY ON ORDER
        for pattern in ['on order', 'ordered', 'incoming']:
            matches = [col for col in source_cols.keys() if pattern in col]
            if matches:
                mappings['QTY ON ORDER'] = self.df_source[source_cols[matches[0]]]
                print(f"QTY ON ORDER: {matches[0]}")
                break

        # Build the master dataframe
        num_rows = len(self.df_source)

        for col in MASTER_COLUMNS:
            if col in mappings:
                if isinstance(mappings[col], str):
                    # Static value (like supplier name)
                    self.df_master[col] = [mappings[col]] * num_rows
                else:
                    # Column data
                    self.df_master[col] = mappings[col].values
            else:
                # Empty column
                self.df_master[col] = [None] * num_rows
                if col not in ['Supplier Code', 'Brand Sub Tag', 'NEXT SHIPMENT', 'Tags', 'LINKS']:
                    self.issues.append(f"Column '{col}' not mapped")

        print(f"\nMaster dataframe created: {self.df_master.shape}")
        return True

    def validate_data(self):
        """Data quality validation"""
        print(f"\nData Quality Check:")
        print("-" * 80)

        # Check required fields
        required = ['SKU / MODEL ', 'PRODUCT DESCRIPTION']
        for col in required:
            null_count = self.df_master[col].isna().sum()
            if null_count > 0:
                pct = (null_count / len(self.df_master)) * 100
                print(f"⚠ {col}: {null_count} null values ({pct:.1f}%)")
                if pct > 50:
                    self.issues.append(f"{col} is mostly empty ({pct:.1f}%)")

        # Check numeric fields
        numeric_cols = ['SUPPLIER SOH', 'COST  EX VAT', 'QTY ON ORDER']
        for col in numeric_cols:
            if self.df_master[col].notna().any():
                try:
                    # Try to convert to numeric
                    self.df_master[col] = pd.to_numeric(self.df_master[col], errors='coerce')
                    print(f"✓ {col}: converted to numeric")
                except:
                    print(f"⚠ {col}: could not convert to numeric")

        return True

    def save_to_master(self):
        """Save to Consolidated_Supplier_Data.xlsx as new sheet"""
        print(f"\nSaving to Master File:")
        print("-" * 80)

        try:
            # Load existing workbook
            wb = load_workbook(MASTER_FILE)

            # Create sheet name (max 31 chars for Excel)
            sheet_name = self.supplier_name[:31]

            # Remove sheet if it exists
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
                print(f"Removed existing sheet '{sheet_name}'")

            # Create new sheet
            ws = wb.create_sheet(sheet_name)

            # Write headers
            for col_idx, col_name in enumerate(MASTER_COLUMNS, 1):
                ws.cell(row=1, column=col_idx, value=col_name)

            # Write data
            for row_idx, row in enumerate(self.df_master.itertuples(index=False), 2):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # Save workbook
            wb.save(MASTER_FILE)
            print(f"✓ Saved {len(self.df_master)} rows to sheet '{sheet_name}'")

            return True

        except Exception as e:
            self.issues.append(f"Save error: {str(e)}")
            print(f"ERROR: {e}")
            return False

    def process(self):
        """Complete processing workflow"""
        if not self.analyze_file():
            return False

        if not self.map_columns():
            return False

        if not self.validate_data():
            return False

        if not self.save_to_master():
            return False

        return True


def main():
    """Process all batch 2 files"""
    print("\n" + "="*80)
    print("SUPPLIER DATA CONSOLIDATION - BATCH 2")
    print("="*80)

    results = {}

    for filename, supplier_name in BATCH_2_FILES.items():
        source_file = SOURCE_DIR / filename

        if not source_file.exists():
            print(f"\n❌ File not found: {filename}")
            continue

        # Skip very large files for now (>50MB)
        file_size_mb = source_file.stat().st_size / (1024 * 1024)
        if file_size_mb > 50:
            print(f"\n⚠ Skipping large file: {filename} ({file_size_mb:.1f}MB)")
            results[supplier_name] = {
                'status': 'skipped',
                'reason': f'File too large ({file_size_mb:.1f}MB)',
                'rows': 0
            }
            continue

        processor = SupplierProcessor(source_file, supplier_name)
        success = processor.process()

        results[supplier_name] = {
            'status': 'success' if success else 'failed',
            'rows': len(processor.df_master) if processor.df_master is not None else 0,
            'issues': processor.issues
        }

    # Summary Report
    print("\n" + "="*80)
    print("BATCH 2 PROCESSING SUMMARY")
    print("="*80)

    for supplier, result in results.items():
        print(f"\n{supplier}:")
        print(f"  Status: {result['status']}")
        print(f"  Rows: {result['rows']}")
        if result.get('issues'):
            print(f"  Issues:")
            for issue in result['issues']:
                print(f"    - {issue}")
        if result.get('reason'):
            print(f"  Reason: {result['reason']}")

    # Overall stats
    total_rows = sum(r['rows'] for r in results.values())
    success_count = sum(1 for r in results.values() if r['status'] == 'success')

    print(f"\n{'='*80}")
    print(f"OVERALL: {success_count}/{len(BATCH_2_FILES)} files processed successfully")
    print(f"Total rows added: {total_rows:,}")
    print(f"{'='*80}\n")


if __name__ == '__main__':
    main()
