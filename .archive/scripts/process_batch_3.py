#!/usr/bin/env python3
"""
DATA ORACLE AGENT ALPHA-3: Batch 3 Supplier Data Consolidation
Processes 7 supplier files into Master format (13 columns)
"""

import pandas as pd
import openpyxl
from pathlib import Path
import re
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

# Master template columns
MASTER_COLUMNS = [
    'Supplier Name', 'Supplier Code', 'Product Category', 'BRAND',
    'Brand Sub Tag', 'SKU / MODEL', 'PRODUCT DESCRIPTION', 'SUPPLIER SOH',
    'COST EX VAT', 'QTY ON ORDER', 'NEXT SHIPMENT', 'Tags', 'LINKS'
]

# File definitions
BATCH3_FILES = {
    'SonicInformed_20250808_I.xlsx': {
        'supplier_name': 'Sonic Informed',
        'supplier_code': 'SONIC-INF'
    },
    'Stage Audio Works SOH info_20250826_0247.xlsx': {
        'supplier_name': 'Stage Audio Works',
        'supplier_code': 'STAGE-AUDIO'
    },
    'Stage one FullPL-DP (7).xlsx': {
        'supplier_name': 'Stage One Distribution',
        'supplier_code': 'STAGE-ONE'
    },
    'Tuerk Multimedia  Studio Price List_ JULY 2025.xlsx': {
        'supplier_name': 'Tuerk Multimedia',
        'supplier_code': 'TUERK-MULTI'
    },
    'Tuerk Tech General Price List JULY 2025.xls': {
        'supplier_name': 'Tuerk Technologies',
        'supplier_code': 'TUERK-TECH'
    },
    'Viva Afrika Dealer Price List 07 May 2025.xlsx': {
        'supplier_name': 'Viva Afrika',
        'supplier_code': 'VIVA-AFRIKA'
    },
    'YAMAHA RETAIL PRICELIST NOV 2024 (2).xlsx': {
        'supplier_name': 'Yamaha',
        'supplier_code': 'YAMAHA'
    }
}

SOURCE_DIR = Path('/mnt/k/00Project/MantisNXT/database/Uploads/drive-download-20250904T012253Z-1-001')
MASTER_FILE = Path('/mnt/k/00Project/MantisNXT/database/Uploads/Consolidated_Batch3_Data.xlsx')

def clean_numeric(value):
    """Convert value to float"""
    if pd.isna(value):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        value = re.sub(r'[R$€£,\s]', '', value)
        try:
            return float(value)
        except:
            return 0.0
    return 0.0

def standardize_brand(brand):
    """Standardize brand names"""
    if pd.isna(brand) or str(brand).strip() == '':
        return 'Unknown'
    
    brand = str(brand).strip().upper()
    brand_map = {
        'SHURE': 'Shure', 'SENNHEISER': 'Sennheiser', 'YAMAHA': 'Yamaha',
        'RODE': 'Rode', 'AUDIO-TECHNICA': 'Audio-Technica', 'MACKIE': 'Mackie',
        'BEHRINGER': 'Behringer', 'FOCUSRITE': 'Focusrite', 'JBL': 'JBL',
        'QSC': 'QSC', 'ALLEN & HEATH': 'Allen & Heath'
    }
    return brand_map.get(brand, brand.title())

def analyze_structure(file_path):
    """Analyze file structure"""
    print(f"\n{'='*80}")
    print(f"ANALYZING: {file_path.name}")
    print(f"{'='*80}")
    
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        print(f"Sheets: {wb.sheetnames}")
        
        for sheet_name in wb.sheetnames:
            print(f"\n--- Sheet: {sheet_name} ---")
            try:
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=10)
                print(f"Dimensions: {df.shape[0]} rows × {df.shape[1]} columns")
                print(df.head(5).to_string())
            except Exception as e:
                print(f"Error: {e}")
        wb.close()
    except Exception as e:
        print(f"ERROR: {e}")

def process_supplier(file_path, supplier_info):
    """Generic processor for supplier files"""
    print(f"\n{'='*80}")
    print(f"PROCESSING: {supplier_info['supplier_name']}")
    print(f"{'='*80}")
    
    try:
        # Try to find header row
        df = None
        for header_row in range(0, 10):
            try:
                temp_df = pd.read_excel(file_path, header=header_row)
                if len(temp_df.columns) > 3 and len(temp_df) > 5:
                    df = temp_df
                    print(f"Found data at row {header_row}")
                    print(f"Columns: {list(df.columns)[:10]}")
                    break
            except:
                continue
        
        if df is None:
            print("Could not find valid data")
            return None
        
        # Map columns
        col_map = {}
        for col in df.columns:
            col_lower = str(col).lower()
            if 'brand' in col_lower or 'make' in col_lower:
                col_map['brand'] = col
            elif 'model' in col_lower or 'sku' in col_lower or 'code' in col_lower or 'part' in col_lower:
                col_map['sku'] = col
            elif 'description' in col_lower or 'product' in col_lower or 'name' in col_lower:
                col_map['description'] = col
            elif 'stock' in col_lower or 'soh' in col_lower or 'qty' in col_lower or 'quantity' in col_lower:
                col_map['stock'] = col
            elif 'price' in col_lower or 'cost' in col_lower or 'dealer' in col_lower or 'rrp' in col_lower:
                col_map['price'] = col
            elif 'category' in col_lower or 'type' in col_lower or 'group' in col_lower:
                col_map['category'] = col
        
        print(f"Column mappings: {col_map}")
        
        # Create master DataFrame
        master_df = pd.DataFrame(columns=MASTER_COLUMNS)
        
        master_df['Supplier Name'] = supplier_info['supplier_name']
        master_df['Supplier Code'] = supplier_info['supplier_code']
        master_df['Product Category'] = df[col_map['category']].astype(str) if 'category' in col_map else 'General'
        master_df['BRAND'] = df[col_map['brand']].apply(standardize_brand) if 'brand' in col_map else 'Unknown'
        master_df['Brand Sub Tag'] = ''
        master_df['SKU / MODEL'] = df[col_map['sku']].astype(str) if 'sku' in col_map else ''
        master_df['PRODUCT DESCRIPTION'] = df[col_map['description']].astype(str) if 'description' in col_map else ''
        master_df['SUPPLIER SOH'] = df[col_map['stock']].apply(clean_numeric) if 'stock' in col_map else 0
        master_df['COST EX VAT'] = df[col_map['price']].apply(clean_numeric) if 'price' in col_map else 0.0
        master_df['QTY ON ORDER'] = 0
        master_df['NEXT SHIPMENT'] = ''
        master_df['Tags'] = ''
        master_df['LINKS'] = ''
        
        # Remove empty rows
        master_df = master_df[master_df['SKU / MODEL'].str.strip() != '']
        
        print(f"Processed {len(master_df)} rows")
        return master_df
        
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()
        return None

def main():
    print(f"""
{'#'*80}
# DATA ORACLE AGENT ALPHA-3: Batch 3 Supplier Consolidation
# Processing 7 supplier files
{'#'*80}
""")
    
    # Phase 1: Analyze
    print("\nPHASE 1: ANALYSIS")
    for filename in BATCH3_FILES.keys():
        file_path = SOURCE_DIR / filename
        if file_path.exists():
            analyze_structure(file_path)
    
    # Phase 2: Process
    print("\nPHASE 2: PROCESSING")
    results = {}
    
    for filename, supplier_info in BATCH3_FILES.items():
        file_path = SOURCE_DIR / filename
        if not file_path.exists():
            print(f"SKIP: {filename} not found")
            continue
        
        df = process_supplier(file_path, supplier_info)
        if df is not None and len(df) > 0:
            results[supplier_info['supplier_name']] = df
    
    # Phase 3: Write
    print("\nPHASE 3: WRITING TO EXCEL")
    if not results:
        print("ERROR: No data to write")
        return
    
    try:
        with pd.ExcelWriter(MASTER_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            for supplier_name, df in results.items():
                sheet_name = supplier_name[:31]
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                print(f"✓ Created '{sheet_name}' with {len(df)} rows")
        
        # Phase 4: Report
        print("\nPHASE 4: VALIDATION REPORT")
        total_rows = 0
        for supplier_name, df in results.items():
            print(f"\n{supplier_name}:")
            print(f"  Rows: {len(df)}")
            print(f"  Unique SKUs: {df['SKU / MODEL'].nunique()}")
            print(f"  Brands: {df['BRAND'].nunique()}")
            print(f"  Stock Value: R{(df['SUPPLIER SOH'] * df['COST EX VAT']).sum():,.2f}")
            total_rows += len(df)
        
        print(f"\nTOTAL: {total_rows} rows across {len(results)} suppliers")
        
    except Exception as e:
        print(f"ERROR writing: {e}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    main()
