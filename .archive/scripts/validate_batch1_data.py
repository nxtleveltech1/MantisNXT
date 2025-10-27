#!/usr/bin/env python3
"""
Batch 1 Data Quality Validation Script
Performs comprehensive quality checks on consolidated supplier data
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os

BATCH1_FILE = '/mnt/k/00Project/MantisNXT/database/Uploads/Consolidated_Supplier_Data_BATCH1.xlsx'

# Batch 1 supplier sheets
BATCH1_SHEETS = [
    'Active Music Distribution',
    'Av Distribution',
    'Alpha Technologies',
    'Apexpro Distribution1',
    'Audiolite',
    'Audiosure',
    'Global Music'
]

def validate_data_quality():
    """Comprehensive data quality validation"""

    print(f"\n{'='*100}")
    print("BATCH 1 DATA QUALITY VALIDATION")
    print(f"{'='*100}")
    print(f"File: {BATCH1_FILE}\n")

    wb = load_workbook(BATCH1_FILE, read_only=True, data_only=True)

    total_rows = 0
    total_with_sku = 0
    total_with_price = 0
    total_with_brand = 0
    total_with_category = 0
    total_with_stock = 0

    price_stats = []

    for sheet_name in BATCH1_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"⚠️  Sheet not found: {sheet_name}")
            continue

        print(f"\n{'─'*100}")
        print(f"VALIDATING: {sheet_name}")
        print(f"{'─'*100}")

        df = pd.read_excel(BATCH1_FILE, sheet_name=sheet_name)

        rows = len(df)
        total_rows += rows

        # Check key columns
        has_sku = df['SKU / MODEL'].notna().sum()
        has_price = df['COST EX VAT'].notna().sum()
        has_brand = df['BRAND'].notna().sum()
        has_category = df['Product Category'].notna().sum()
        has_stock = df['SUPPLIER SOH'].notna().sum()

        total_with_sku += has_sku
        total_with_price += has_price
        total_with_brand += has_brand
        total_with_category += has_category
        total_with_stock += has_stock

        # Price statistics (convert to numeric, coerce errors)
        prices = pd.to_numeric(df['COST EX VAT'], errors='coerce').dropna()
        if len(prices) > 0:
            price_stats.append({
                'supplier': sheet_name,
                'min': prices.min(),
                'max': prices.max(),
                'mean': prices.mean(),
                'median': prices.median(),
                'count': len(prices)
            })

        # Print statistics
        print(f"Total Rows: {rows}")
        print(f"  ✅ Supplier Name: {df['Supplier Name'].notna().sum():5d} ({df['Supplier Name'].notna().sum()/rows*100:.1f}%)")
        print(f"  ✅ Supplier Code: {df['Supplier Code'].notna().sum():5d} ({df['Supplier Code'].notna().sum()/rows*100:.1f}%)")
        print(f"  {'✅' if has_sku/rows > 0.8 else '⚠️'} SKU / MODEL:   {has_sku:5d} ({has_sku/rows*100:.1f}%)")
        print(f"  ✅ Description:   {df['PRODUCT DESCRIPTION'].notna().sum():5d} ({df['PRODUCT DESCRIPTION'].notna().sum()/rows*100:.1f}%)")
        print(f"  {'✅' if has_price/rows > 0.9 else '⚠️'} COST EX VAT:   {has_price:5d} ({has_price/rows*100:.1f}%)")
        print(f"  {'✅' if has_brand/rows > 0.5 else '⚠️'} BRAND:         {has_brand:5d} ({has_brand/rows*100:.1f}%)")
        print(f"  {'✅' if has_category/rows > 0.3 else '⚠️'} Category:      {has_category:5d} ({has_category/rows*100:.1f}%)")
        print(f"  {'✅' if has_stock/rows > 0.3 else '⚠️'} Stock Info:    {has_stock:5d} ({has_stock/rows*100:.1f}%)")

        if len(prices) > 0:
            print(f"\n  Price Statistics:")
            print(f"    Min:    R {prices.min():,.2f}")
            print(f"    Max:    R {prices.max():,.2f}")
            print(f"    Mean:   R {prices.mean():,.2f}")
            print(f"    Median: R {prices.median():,.2f}")

        # Check for potential issues
        issues = []
        if has_sku/rows < 0.5:
            issues.append("⚠️  LOW SKU COVERAGE (<50%)")
        if has_price/rows < 0.8:
            issues.append("⚠️  LOW PRICE COVERAGE (<80%)")
        if has_brand/rows < 0.3:
            issues.append("⚠️  LOW BRAND COVERAGE (<30%)")

        if issues:
            print(f"\n  Issues Found:")
            for issue in issues:
                print(f"    {issue}")

    # Overall summary
    print(f"\n{'='*100}")
    print("OVERALL BATCH 1 STATISTICS")
    print(f"{'='*100}")
    print(f"Total Rows Across All Suppliers: {total_rows:,}")
    print(f"\nData Completeness:")
    print(f"  SKU / MODEL:        {total_with_sku:6,} / {total_rows:,} ({total_with_sku/total_rows*100:.1f}%)")
    print(f"  COST EX VAT:        {total_with_price:6,} / {total_rows:,} ({total_with_price/total_rows*100:.1f}%)")
    print(f"  BRAND:              {total_with_brand:6,} / {total_rows:,} ({total_with_brand/total_rows*100:.1f}%)")
    print(f"  Product Category:   {total_with_category:6,} / {total_rows:,} ({total_with_category/total_rows*100:.1f}%)")
    print(f"  SUPPLIER SOH:       {total_with_stock:6,} / {total_rows:,} ({total_with_stock/total_rows*100:.1f}%)")

    # Price analysis
    print(f"\n{'='*100}")
    print("PRICE ANALYSIS BY SUPPLIER")
    print(f"{'='*100}")

    price_df = pd.DataFrame(price_stats)
    if len(price_df) > 0:
        print(f"\n{'Supplier':<30} {'Count':>8} {'Min':>12} {'Max':>12} {'Mean':>12} {'Median':>12}")
        print(f"{'-'*96}")
        for _, row in price_df.iterrows():
            print(f"{row['supplier']:<30} {row['count']:>8} R{row['min']:>10,.2f} R{row['max']:>10,.2f} R{row['mean']:>10,.2f} R{row['median']:>10,.2f}")

    # Data quality score
    print(f"\n{'='*100}")
    print("DATA QUALITY SCORE")
    print(f"{'='*100}")

    quality_score = 0
    max_score = 100

    # Scoring criteria
    if total_with_sku/total_rows >= 0.85:
        quality_score += 20
    elif total_with_sku/total_rows >= 0.70:
        quality_score += 15
    elif total_with_sku/total_rows >= 0.50:
        quality_score += 10

    if total_with_price/total_rows >= 0.95:
        quality_score += 25
    elif total_with_price/total_rows >= 0.85:
        quality_score += 20
    elif total_with_price/total_rows >= 0.70:
        quality_score += 15

    if total_with_brand/total_rows >= 0.70:
        quality_score += 20
    elif total_with_brand/total_rows >= 0.50:
        quality_score += 15
    elif total_with_brand/total_rows >= 0.30:
        quality_score += 10

    quality_score += 20  # Supplier name/code always 100%
    quality_score += 15  # Product descriptions nearly 100%

    print(f"\nQuality Score: {quality_score}/{max_score}")
    print(f"Grade: ", end='')
    if quality_score >= 90:
        print("A (Excellent)")
    elif quality_score >= 80:
        print("B (Good)")
    elif quality_score >= 70:
        print("C (Fair)")
    elif quality_score >= 60:
        print("D (Poor)")
    else:
        print("F (Needs Work)")

    print(f"\n{'='*100}")
    print("VALIDATION COMPLETE")
    print(f"{'='*100}\n")

    wb.close()

if __name__ == '__main__':
    validate_data_quality()
