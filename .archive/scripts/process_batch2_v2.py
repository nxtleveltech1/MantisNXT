#!/usr/bin/env python3
"""
Supplier Data Consolidation - Batch 2 Processor v2
Improved header detection and data extraction
"""

import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
import warnings
import re
warnings.filterwarnings('ignore')

SOURCE_DIR = Path('/mnt/k/00Project/MantisNXT/database/Uploads/drive-download-20250904T012253Z-1-001')
MASTER_FILE = Path('/mnt/k/00Project/MantisNXT/database/Uploads/Consolidated_Supplier_Data_Batch2.xlsx')

MASTER_COLUMNS = [
    'Supplier Name ', 'Supplier Code', 'Produt Category', 'BRAND', 'Brand Sub Tag',
    'SKU / MODEL ', 'PRODUCT DESCRIPTION', 'SUPPLIER SOH', 'COST  EX VAT',
    'QTY ON ORDER', 'NEXT SHIPMENT', 'Tags', 'LINKS'
]

def clean_text(text):
    """Clean and normalize text"""
    if pd.isna(text):
        return None
    text = str(text).strip()
    return text if text and text.lower() not in ['nan', 'none', ''] else None

def find_header_row(df):
    """Find the row that contains column headers"""
    for idx in range(min(10, len(df))):
        row = df.iloc[idx]
        # Check if row contains header-like text
        text_count = sum(1 for val in row if isinstance(val, str) and len(str(val)) > 2)
        if text_count >= 3:  # At least 3 text columns
            return idx
    return 0

def process_md_distribution():
    """MD External Stock 2025-08-25.xlsx"""
    supplier = 'MD Distribution'
    print(f"\n{'='*80}\nProcessing: {supplier}\n{'='*80}")

    filepath = SOURCE_DIR / 'MD External Stock 2025-08-25.xlsx'
    df = pd.read_excel(filepath, sheet_name='Sheet1')

    # Headers are in row 0
    df.columns = df.iloc[0]
    df = df[1:].reset_index(drop=True)

    result = pd.DataFrame(columns=MASTER_COLUMNS)
    result['Supplier Name '] = supplier
    result['BRAND'] = df['Brand'].apply(clean_text)
    result['SKU / MODEL '] = df['Item Code'].apply(clean_text)
    result['PRODUCT DESCRIPTION'] = df['Description'].apply(clean_text)
    result['COST  EX VAT'] = pd.to_numeric(df['List Price'], errors='coerce')

    print(f"✓ Extracted {len(result)} rows")
    return supplier, result

def process_music_power():
    """Music Power Pricelist (August 2025).xlsx"""
    supplier = 'Music Power'
    print(f"\n{'='*80}\nProcessing: {supplier}\n{'='*80}")

    filepath = SOURCE_DIR / 'Music Power Pricelist (August 2025).xlsx'

    # Use the 'Order Sheet' which has all products
    df = pd.read_excel(filepath, sheet_name='Order Sheet')

    # Headers are in row 0
    df.columns = df.iloc[0]
    df = df[1:].reset_index(drop=True)

    result = pd.DataFrame(columns=MASTER_COLUMNS)
    result['Supplier Name '] = supplier
    result['SKU / MODEL '] = df['Code'].apply(clean_text)
    result['PRODUCT DESCRIPTION'] = df['Description'].apply(clean_text)
    result['COST  EX VAT'] = pd.to_numeric(df['RETAIL INCL VAT'], errors='coerce')
    result['SUPPLIER SOH'] = pd.to_numeric(df['SOH'], errors='coerce')

    # Extract brand from description if possible
    def extract_brand(desc):
        if pd.isna(desc):
            return None
        desc = str(desc).upper()
        # Common brands in the file
        brands = ['AQUARIAN', 'ARTURIA', 'DIXON', 'DW', 'ELIXIR', 'FZONE',
                 'GIBRALTAR', 'GATOR', 'GEWA', 'HALIFAX', 'SAGA', 'WARWICK']
        for brand in brands:
            if brand in desc:
                return brand
        return None

    result['BRAND'] = df['Description'].apply(extract_brand)

    print(f"✓ Extracted {len(result)} rows")
    return supplier, result

def process_planetworld():
    """Planetworld SOH 20 June.xlsx"""
    supplier = 'Planetworld'
    print(f"\n{'='*80}\nProcessing: {supplier}\n{'='*80}")

    filepath = SOURCE_DIR / 'Planetworld SOH 20 June.xlsx'
    df = pd.read_excel(filepath, sheet_name='Sheet1')

    result = pd.DataFrame(columns=MASTER_COLUMNS)
    result['Supplier Name '] = supplier
    result['SKU / MODEL '] = df['Item No.'].apply(clean_text)
    result['PRODUCT DESCRIPTION'] = df['Item Description'].apply(clean_text)
    result['SUPPLIER SOH'] = pd.to_numeric(df['SOH'], errors='coerce')
    result['Produt Category'] = df['Product Division'].apply(clean_text)

    # Extract brand from description
    def extract_brand(desc):
        if pd.isna(desc):
            return None
        # First word is often the brand
        words = str(desc).split()
        if words:
            return words[0].strip()
        return None

    result['BRAND'] = df['Item Description'].apply(extract_brand)

    print(f"✓ Extracted {len(result)} rows")
    return supplier, result

def process_rockit():
    """Rockit Price_List_25.08.2025.01.xlsx"""
    supplier = 'Rockit'
    print(f"\n{'='*80}\nProcessing: {supplier}\n{'='*80}")

    filepath = SOURCE_DIR / 'Rockit Price_List_25.08.2025.01.xlsx'
    df = pd.read_excel(filepath, sheet_name='Sheet1')

    # Headers are in row 0
    df.columns = df.iloc[0]
    df = df[1:].reset_index(drop=True)

    result = pd.DataFrame(columns=MASTER_COLUMNS)
    result['Supplier Name '] = supplier
    result['SKU / MODEL '] = df.iloc[:, 0].apply(clean_text)  # First column
    result['PRODUCT DESCRIPTION'] = df.iloc[:, 1].apply(clean_text)  # Second column

    # Price might be in column 2 or 3
    for col_idx in range(2, min(4, len(df.columns))):
        col_data = pd.to_numeric(df.iloc[:, col_idx], errors='coerce')
        if col_data.notna().sum() > 10:  # Has numeric data
            result['COST  EX VAT'] = col_data
            break

    print(f"✓ Extracted {len(result)} rows")
    return supplier, result

def process_rolling_thunder():
    """Rolling Thunder July 2025 Pricelist .xlsx"""
    supplier = 'Rolling Thunder'
    print(f"\n{'='*80}\nProcessing: {supplier}\n{'='*80}")

    filepath = SOURCE_DIR / 'Rolling Thunder July 2025 Pricelist .xlsx'

    # Has multiple brand sheets - combine them all
    xl = pd.ExcelFile(filepath)
    all_data = []

    for sheet in xl.sheet_names:
        df = pd.read_excel(filepath, sheet_name=sheet)
        if len(df) > 5:  # Skip empty sheets
            # Headers usually in first row
            df.columns = df.iloc[0]
            df = df[1:].reset_index(drop=True)

            sheet_data = pd.DataFrame(columns=MASTER_COLUMNS)
            sheet_data['Supplier Name '] = supplier
            sheet_data['BRAND'] = sheet  # Sheet name is the brand
            sheet_data['SKU / MODEL '] = df.get('Item Name', df.get('Item Code', None))
            sheet_data['PRODUCT DESCRIPTION'] = df.get('Item Description', df.get('Description', None))
            sheet_data['COST  EX VAT'] = pd.to_numeric(
                df.get('Dealer Excluding VAT', df.get('Dealer Ex VAT', None)),
                errors='coerce'
            )

            all_data.append(sheet_data)

    result = pd.concat(all_data, ignore_index=True) if all_data else pd.DataFrame(columns=MASTER_COLUMNS)
    print(f"✓ Extracted {len(result)} rows from {len(all_data)} sheets")
    return supplier, result

def process_sennheiser():
    """Sennheiser 2025 (2).xlsx"""
    supplier = 'Sennheiser'
    print(f"\n{'='*80}\nProcessing: {supplier}\n{'='*80}")

    filepath = SOURCE_DIR / 'Sennheiser 2025 (2).xlsx'
    df = pd.read_excel(filepath, sheet_name='17th June Full Run Out')

    result = pd.DataFrame(columns=MASTER_COLUMNS)
    result['Supplier Name '] = supplier
    result['BRAND'] = 'Sennheiser'
    result['SKU / MODEL '] = df['Item No.'].apply(clean_text)
    result['PRODUCT DESCRIPTION'] = df['Item Description'].apply(clean_text)
    result['COST  EX VAT'] = pd.to_numeric(df['Retail Ex Vat'], errors='coerce')

    print(f"✓ Extracted {len(result)} rows")
    return supplier, result

def save_to_excel(supplier_name, df):
    """Save supplier data to Excel file"""
    try:
        wb = load_workbook(MASTER_FILE)

        # Sheet name max 31 chars
        sheet_name = supplier_name[:31]

        # Remove if exists
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

        ws = wb.create_sheet(sheet_name)

        # Write headers
        for col_idx, col_name in enumerate(MASTER_COLUMNS, 1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # Write data
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        wb.save(MASTER_FILE)
        print(f"✓ Saved to sheet '{sheet_name}'")
        return True

    except Exception as e:
        print(f"✗ Save error: {e}")
        return False

def main():
    """Process all suppliers"""
    print("\n" + "="*80)
    print("SUPPLIER DATA CONSOLIDATION - BATCH 2 v2")
    print("="*80)

    processors = [
        process_md_distribution,
        process_music_power,
        process_planetworld,
        process_rockit,
        process_rolling_thunder,
        process_sennheiser
    ]

    results = {}

    for processor in processors:
        try:
            supplier, df = processor()

            # Remove completely empty rows
            df = df.dropna(how='all')

            # Save to Excel
            success = save_to_excel(supplier, df)

            results[supplier] = {
                'status': 'success' if success else 'failed',
                'rows': len(df),
                'has_sku': df['SKU / MODEL '].notna().sum(),
                'has_desc': df['PRODUCT DESCRIPTION'].notna().sum(),
                'has_price': df['COST  EX VAT'].notna().sum(),
                'has_stock': df['SUPPLIER SOH'].notna().sum()
            }

        except Exception as e:
            print(f"✗ Error processing: {e}")
            import traceback
            traceback.print_exc()
            results[supplier if 'supplier' in locals() else 'Unknown'] = {
                'status': 'error',
                'error': str(e)
            }

    # Summary
    print("\n" + "="*80)
    print("BATCH 2 SUMMARY")
    print("="*80)

    for supplier, info in results.items():
        print(f"\n{supplier}:")
        print(f"  Status: {info['status']}")
        if info['status'] == 'success':
            print(f"  Rows: {info['rows']}")
            print(f"  SKUs: {info['has_sku']}")
            print(f"  Descriptions: {info['has_desc']}")
            print(f"  Prices: {info['has_price']}")
            print(f"  Stock: {info['has_stock']}")
        elif 'error' in info:
            print(f"  Error: {info['error']}")

    total_rows = sum(r['rows'] for r in results.values() if 'rows' in r)
    success_count = sum(1 for r in results.values() if r['status'] == 'success')

    print(f"\n{'='*80}")
    print(f"Total: {success_count}/{len(processors)} suppliers processed")
    print(f"Total rows: {total_rows:,}")
    print(f"{'='*80}\n")

if __name__ == '__main__':
    main()
