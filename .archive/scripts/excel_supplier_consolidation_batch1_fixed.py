#!/usr/bin/env python3
"""
Excel Supplier Data Consolidation - Batch 1 FIXED
Handles each supplier's unique data structure
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import re

# Master template columns
MASTER_COLUMNS = [
    'Supplier Name', 'Supplier Code', 'Product Category', 'BRAND',
    'Brand Sub Tag', 'SKU / MODEL', 'PRODUCT DESCRIPTION',
    'SUPPLIER SOH', 'COST EX VAT', 'QTY ON ORDER', 'NEXT SHIPMENT',
    'Tags', 'LINKS'
]

# File paths
MASTER_FILE_INPUT = '/mnt/k/00Project/MantisNXT/database/Uploads/Consolidated_Supplier_Data.xlsx'
MASTER_FILE_OUTPUT = '/mnt/k/00Project/MantisNXT/database/Uploads/Consolidated_Supplier_Data_BATCH1.xlsx'
SOURCE_DIR = '/mnt/k/00Project/MantisNXT/database/Uploads/drive-download-20250904T012253Z-1-001/'

def extract_supplier_name(filename):
    """Extract clean supplier name"""
    name = filename.replace('.xlsx', '').replace('.xls', '')
    noise_patterns = [
        r'\s*-?\s*pricelist.*$', r'\s*-?\s*price\s*list.*$',
        r'\s*-?\s*(january|february|march|april|may|june|july|august|september|october|november|december)\s*\d{4}.*$',
        r'\s*-?\s*\d{2}-\d{2}-\d{4}.*$', r'\s*-?\s*\d{4}.*$',
        r'\s*-?\s*stockfile.*$', r'\s*-?\s*stock\s*file.*$',
        r'\s*-?\s*SOH.*$', r'\s*-?\s*\(\d+\)\s*$',
        r'\s*-?\s*fullp.*$', r'\s*-?\s*suggested\s*online.*$',
        r'\s*-?\s*advertised.*$', r'\s*-?\s*sku\s*list.*$',
        r'\s*-?\s*format.*$', r'\s*-\s*$', r'\s+$'
    ]
    for pattern in noise_patterns:
        name = re.sub(pattern, '', name, flags=re.IGNORECASE)
    name = re.sub(r'\s+', ' ', name)
    name = re.sub(r'-+', ' ', name)
    name = name.strip(' -')
    name = ' '.join(word.capitalize() for word in name.split())
    return name

def generate_supplier_code(supplier_name, row_num):
    """Generate supplier code"""
    code_parts = ''.join([word[0].upper() for word in supplier_name.split() if word])
    return f"{code_parts}-{row_num:04d}"

# ====================================================================================
# SUPPLIER-SPECIFIC PROCESSORS
# ====================================================================================

def process_active_music(filepath, supplier_name):
    """Active Music Distribution - Single column price list"""
    print(f"Processing {supplier_name} with custom handler...")

    df = pd.read_excel(filepath, sheet_name='August Pricelist v3', header=None)

    # Extract data - appears to be in first column after header
    rows = []
    for idx, row in df.iterrows():
        if idx == 0:  # Skip header
            continue

        # Try to parse the row
        desc = str(row[0]) if pd.notna(row[0]) else ''
        price = row[1] if len(row) > 1 and pd.notna(row[1]) else None

        if desc and desc != 'nan':
            rows.append({
                'Supplier Name': supplier_name,
                'Supplier Code': generate_supplier_code(supplier_name, len(rows) + 1),
                'PRODUCT DESCRIPTION': desc,
                'COST EX VAT': price
            })

    result_df = pd.DataFrame(rows, columns=MASTER_COLUMNS)
    return result_df

def process_av_distribution(filepath, supplier_name):
    """AV Distribution - Multi-brand sheets"""
    print(f"Processing {supplier_name} with custom handler...")

    all_rows = []
    wb = load_workbook(filepath, read_only=True, data_only=True)

    # Process each brand sheet
    for sheet_name in wb.sheetnames:
        if sheet_name in ['Disclaimer ', 'Front Page']:
            continue

        df = pd.read_excel(filepath, sheet_name=sheet_name)

        # Find header row - look for "SAP Item Code" or similar
        header_row = None
        for idx, row in df.iterrows():
            if 'SAP Item Code' in str(row.values) or 'Description' in str(row.values):
                header_row = idx
                break

        if header_row is None:
            continue

        # Re-read with correct header
        df = pd.read_excel(filepath, sheet_name=sheet_name, header=header_row)

        # Extract data
        for idx, row in df.iterrows():
            sku = str(row.get('SAP Item Code', '')) if 'SAP Item Code' in df.columns else ''
            desc = str(row.get('Description', '')) if 'Description' in df.columns else ''
            price = row.get('Retail Price (Ex VAT)', None) if 'Retail Price (Ex VAT)' in df.columns else None

            # Skip if essential data missing
            if sku == 'nan' or desc == 'nan':
                continue

            all_rows.append({
                'Supplier Name': supplier_name,
                'Supplier Code': generate_supplier_code(supplier_name, len(all_rows) + 1),
                'BRAND': sheet_name,  # Sheet name is the brand
                'SKU / MODEL': sku,
                'PRODUCT DESCRIPTION': desc,
                'COST EX VAT': price
            })

    wb.close()
    result_df = pd.DataFrame(all_rows, columns=MASTER_COLUMNS)
    return result_df

def process_alpha_technologies(filepath, supplier_name):
    """Alpha Technologies - Multi-brand sheets with consistent format"""
    print(f"Processing {supplier_name} with custom handler...")

    all_rows = []
    wb = load_workbook(filepath, read_only=True, data_only=True)

    # Process each sheet
    for sheet_name in wb.sheetnames:
        if sheet_name in ['Front Page']:
            continue

        # Read the sheet with header in row 1 (0-indexed)
        df = pd.read_excel(filepath, sheet_name=sheet_name, header=1)

        # Extract data - skip rows with category headers
        for idx, row in df.iterrows():
            # Get values from first 3 columns (MODEL, DESCRIPTION, RETAIL)
            model = str(row.iloc[0]) if len(row) > 0 and pd.notna(row.iloc[0]) else ''
            desc = str(row.iloc[1]) if len(row) > 1 and pd.notna(row.iloc[1]) else ''
            price = row.iloc[2] if len(row) > 2 and pd.notna(row.iloc[2]) else None

            # Skip category headers, empty rows, and info rows
            if (model == 'nan' or model == '' or
                'Click on model' in model or
                'EQUIPMENT' in model.upper() or
                'PROCESSORS' in model.upper() or
                'EXCL VAT' in model.upper()):
                continue

            all_rows.append({
                'Supplier Name': supplier_name,
                'Supplier Code': generate_supplier_code(supplier_name, len(all_rows) + 1),
                'BRAND': sheet_name,
                'SKU / MODEL': model,
                'PRODUCT DESCRIPTION': desc,
                'COST EX VAT': price
            })

    wb.close()
    result_df = pd.DataFrame(all_rows, columns=MASTER_COLUMNS)
    return result_df

def process_apexpro(filepath, supplier_name):
    """ApexPro Distribution - Well-structured single sheet"""
    print(f"Processing {supplier_name} with custom handler...")

    df = pd.read_excel(filepath)

    rows = []
    for idx, row in df.iterrows():
        rows.append({
            'Supplier Name': supplier_name,
            'Supplier Code': generate_supplier_code(supplier_name, idx + 1),
            'BRAND': row.get('Brand', ''),
            'SKU / MODEL': row.get('SKU', ''),
            'PRODUCT DESCRIPTION': row.get('Description', ''),
            'SUPPLIER SOH': row.get('Stock Status', ''),
            'COST EX VAT': row.get('Dealer Cost inc VAT', None),
            'NEXT SHIPMENT': row.get('ETA', '')
        })

    result_df = pd.DataFrame(rows, columns=MASTER_COLUMNS)
    return result_df

def process_audiolite(filepath, supplier_name):
    """Audiolite - Headers in row 3"""
    print(f"Processing {supplier_name} with custom handler...")

    df = pd.read_excel(filepath, header=2)  # Headers in row 3 (0-indexed = 2)

    rows = []
    for idx, row in df.iterrows():
        code = str(row.get('CODE', '')) if 'CODE' in df.columns else ''
        desc = str(row.get('DESCRIPTION', '')) if 'DESCRIPTION' in df.columns else ''
        qty = row.get('QTY', '') if 'QTY' in df.columns else ''
        dealer_price = row.get('DEALER INC', None) if 'DEALER INC' in df.columns else None

        # Skip category headers and empty rows
        if code == 'nan' or desc == 'nan' or 'SPEAKERS' in code.upper() or 'AMPLIFIERS' in code.upper():
            continue

        rows.append({
            'Supplier Name': supplier_name,
            'Supplier Code': generate_supplier_code(supplier_name, len(rows) + 1),
            'SKU / MODEL': code,
            'PRODUCT DESCRIPTION': desc,
            'SUPPLIER SOH': qty,
            'COST EX VAT': dealer_price
        })

    result_df = pd.DataFrame(rows, columns=MASTER_COLUMNS)
    return result_df

def process_audiosure(filepath, supplier_name):
    """Audiosure - Well-structured stock file"""
    print(f"Processing {supplier_name} with custom handler...")

    df = pd.read_excel(filepath)

    rows = []
    for idx, row in df.iterrows():
        category = row.get('Category', '')
        item_num = row.get('ItemNumber', '')
        item_desc = row.get('ItemDescription', '')
        status = row.get('ItemStatus', '')
        retail = row.get('Retail Incl.', None)

        rows.append({
            'Supplier Name': supplier_name,
            'Supplier Code': generate_supplier_code(supplier_name, idx + 1),
            'Product Category': category,
            'SKU / MODEL': item_num,
            'PRODUCT DESCRIPTION': item_desc,
            'SUPPLIER SOH': status,
            'COST EX VAT': retail
        })

    result_df = pd.DataFrame(rows, columns=MASTER_COLUMNS)
    return result_df

def process_global_music(filepath, supplier_name):
    """Global Music - Simple price list"""
    print(f"Processing {supplier_name} with custom handler...")

    df = pd.read_excel(filepath, header=None)

    rows = []
    for idx, row in df.iterrows():
        if idx == 0:  # Skip header
            continue

        desc = str(row[1]) if len(row) > 1 and pd.notna(row[1]) else ''
        price = row[2] if len(row) > 2 and pd.notna(row[2]) else None

        if desc and desc != 'nan':
            rows.append({
                'Supplier Name': supplier_name,
                'Supplier Code': generate_supplier_code(supplier_name, len(rows) + 1),
                'PRODUCT DESCRIPTION': desc,
                'COST EX VAT': price
            })

    result_df = pd.DataFrame(rows, columns=MASTER_COLUMNS)
    return result_df

# ====================================================================================
# MAIN PROCESSING
# ====================================================================================

def process_supplier_file(filepath, supplier_name, processor_func):
    """Process supplier file with custom processor"""
    try:
        df = processor_func(filepath, supplier_name)
        print(f"✅ Extracted {len(df)} rows")
        return df
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def main():
    """Main execution"""
    print(f"\n{'#'*100}")
    print("EXCEL SUPPLIER DATA CONSOLIDATION - BATCH 1 FIXED")
    print(f"{'#'*100}")

    # Define files and their processors
    suppliers = [
        ('ACTIVE MUSIC DISTRIBUTION PRICELIST - AUGUST 2025.xlsx', process_active_music),
        ('AV Distribution Pricelist - May 2025 (1).xlsx', process_av_distribution),
        ('Alpha-Technologies-Pricelist-August-2025- (2).xlsx', process_alpha_technologies),
        ('ApexPro Distribution pricelist 25-04-2025 (1).xlsx', process_apexpro),
        ('Audiolite PRICE LIST-WHOLESALE OUT.xlsx', process_audiolite),
        ('Audiosure StockFile 30072025.xlsx', process_audiosure),
        ('GLOBAL MUSIC SUGGESTED ONLINE ADVERTISED PRICELIST NOVEMBER 2024 SKU LIST FORMAT .xlsx', process_global_music)
    ]

    # Load or create master workbook
    if os.path.exists(MASTER_FILE_INPUT):
        print(f"\n✅ Loading existing master workbook...")
        master_wb = load_workbook(MASTER_FILE_INPUT)
    else:
        master_wb = Workbook()
        master_wb.active.title = 'MASTER'

    results = []

    # Process each supplier
    for filename, processor in suppliers:
        filepath = os.path.join(SOURCE_DIR, filename)
        if not os.path.exists(filepath):
            print(f"\n⚠️  File not found: {filename}")
            continue

        supplier_name = extract_supplier_name(filename)

        print(f"\n{'='*100}")
        print(f"PROCESSING: {supplier_name}")
        print(f"File: {filename}")
        print(f"{'='*100}")

        df = process_supplier_file(filepath, supplier_name, processor)

        if df is not None and len(df) > 0:
            # Create sheet in master workbook
            sheet_name = supplier_name[:31]

            if sheet_name in master_wb.sheetnames:
                del master_wb[sheet_name]

            ws = master_wb.create_sheet(sheet_name)

            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(r_idx, c_idx, value)

            results.append({
                'supplier': supplier_name,
                'rows': len(df),
                'sheet': sheet_name
            })

    # Save workbook
    print(f"\n{'='*100}")
    print("SAVING WORKBOOK...")
    master_wb.save(MASTER_FILE_OUTPUT)
    print(f"✅ Saved: {MASTER_FILE_OUTPUT}")

    # Summary
    print(f"\n{'#'*100}")
    print("BATCH 1 CONSOLIDATION COMPLETE")
    print(f"{'#'*100}")

    total_rows = 0
    for result in results:
        print(f"\n{result['supplier']}")
        print(f"  Sheet: {result['sheet']}")
        print(f"  Rows: {result['rows']}")
        total_rows += result['rows']

    print(f"\n{'='*100}")
    print(f"TOTAL ROWS: {total_rows}")
    print(f"{'='*100}")

if __name__ == '__main__':
    main()
