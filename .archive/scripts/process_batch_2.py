#!/usr/bin/env python3
"""
Supplier Data Processing - Batch 2 (Files 8-14)
Process: MD Distribution, Music Power, Planetworld, Pro Audio Platinum,
         Rockit, Rolling Thunder, Sennheiser
"""

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
import re
from datetime import datetime
from typing import Dict, List, Tuple, Optional
import warnings
warnings.filterwarnings('ignore')

# ═══════════════════════════════════════════════════════════════════
# MASTER TEMPLATE SCHEMA
# ═══════════════════════════════════════════════════════════════════

MASTER_COLUMNS = [
    'Brand', 'Category', 'Sub-Category', 'Type', 'Application',
    'SKU', 'Product Description', 'Country Of Origin',
    'Unit Of Measure', 'Carton Quantity', 'Bar Code',
    'Supplier Code', 'Our Item Code', 'Supplier',
    'Cost Price Excl', 'Cost Price Incl', 'Retail Price Incl',
    'Supplier Price Incl', 'QTY On Hand', 'Date Captured'
]

REQUIRED_FIELDS = ['SKU', 'Product Description', 'Supplier', 'Cost Price Excl']

# ═══════════════════════════════════════════════════════════════════
# FILE CONFIGURATIONS
# ═══════════════════════════════════════════════════════════════════

BATCH_2_CONFIGS = {
    'MD External Stock 2025-08-25.xlsx': {
        'supplier': 'MD Distribution',
        'sheet': 0,
        'skip_rows': 0,
        'mapping': {
            'SKU': ['supplier code', 'sku', 'code', 'item code'],
            'Product Description': ['description', 'product', 'item description'],
            'Cost Price Excl': ['cost', 'price', 'ex vat', 'excl'],
            'QTY On Hand': ['qty', 'stock', 'quantity', 'on hand', 'soh']
        }
    },
    'Music Power Pricelist (August 2025).xlsx': {
        'supplier': 'Music Power',
        'sheet': 0,
        'skip_rows': 0,
        'mapping': {
            'SKU': ['sku', 'code', 'product code'],
            'Product Description': ['description', 'product'],
            'Brand': ['brand', 'make', 'manufacturer'],
            'Cost Price Excl': ['dealer', 'cost', 'price ex', 'excl'],
            'Retail Price Incl': ['rrp', 'retail', 'selling price']
        }
    },
    'Planetworld SOH 20 June.xlsx': {
        'supplier': 'Planetworld',
        'sheet': 0,
        'skip_rows': 0,
        'mapping': {
            'SKU': ['stock code', 'sku', 'code'],
            'Product Description': ['description', 'product'],
            'Cost Price Excl': ['cost', 'unit cost', 'price'],
            'QTY On Hand': ['qty on hand', 'stock', 'soh']
        }
    },
    'Pro Audio platinum .xlsx': {
        'supplier': 'Pro Audio Platinum',
        'sheet': 0,
        'skip_rows': 0,
        'mapping': {
            'SKU': ['part no', 'sku', 'code', 'item'],
            'Product Description': ['description', 'product'],
            'Brand': ['brand', 'make'],
            'Cost Price Excl': ['dealer', 'cost', 'price'],
            'Retail Price Incl': ['rrp', 'retail']
        }
    },
    'Rockit Price_List_25.08.2025.01.xlsx': {
        'supplier': 'Rockit',
        'sheet': 0,
        'skip_rows': 0,
        'mapping': {
            'SKU': ['product code', 'sku', 'code'],
            'Product Description': ['description', 'product'],
            'Brand': ['brand'],
            'Cost Price Excl': ['dealer', 'cost', 'price ex'],
            'Retail Price Incl': ['rrp', 'retail']
        }
    },
    'Rolling Thunder July 2025 Pricelist .xlsx': {
        'supplier': 'Rolling Thunder',
        'sheet': 0,
        'skip_rows': 0,
        'mapping': {
            'SKU': ['code', 'sku', 'item code'],
            'Product Description': ['description', 'product'],
            'Cost Price Excl': ['ex vat', 'excl', 'cost', 'price'],
            'Cost Price Incl': ['inc vat', 'incl']
        }
    },
    'Sennheiser 2025 (2).xlsx': {
        'supplier': 'Sennheiser',
        'sheet': 0,
        'skip_rows': 0,
        'mapping': {
            'SKU': ['article', 'sku', 'code', 'product code'],
            'Product Description': ['description', 'product'],
            'Brand': ['brand'],
            'Cost Price Excl': ['dealer', 'cost', 'net'],
            'Retail Price Incl': ['rrp', 'retail', 'list price']
        }
    }
}

# ═══════════════════════════════════════════════════════════════════
# UTILITY FUNCTIONS
# ═══════════════════════════════════════════════════════════════════

def clean_column_name(col: str) -> str:
    """Normalize column names for matching"""
    if pd.isna(col):
        return ''
    return str(col).lower().strip().replace('_', ' ').replace('-', ' ')

def find_header_row(df: pd.DataFrame, mapping: Dict) -> Optional[int]:
    """Intelligently detect header row"""
    for idx in range(min(20, len(df))):
        row_values = [clean_column_name(str(v)) for v in df.iloc[idx].values]

        # Count potential matches
        matches = 0
        for master_col, patterns in mapping.items():
            for pattern in patterns:
                if any(pattern.lower() in val for val in row_values if val):
                    matches += 1
                    break

        if matches >= 3:  # Need at least 3 field matches
            return idx

    return None

def map_columns(df: pd.DataFrame, mapping: Dict) -> Dict[str, str]:
    """Map source columns to master template columns"""
    column_map = {}
    source_cols = {clean_column_name(col): col for col in df.columns}

    for master_col, patterns in mapping.items():
        for pattern in patterns:
            for cleaned, original in source_cols.items():
                if pattern.lower() in cleaned:
                    column_map[original] = master_col
                    break
            if master_col in column_map.values():
                break

    return column_map

def sanitize_value(value, data_type='str'):
    """Clean and standardize data values"""
    if pd.isna(value) or value == '' or value == 'None':
        return None

    if data_type == 'numeric':
        try:
            # Remove currency symbols, commas, spaces
            if isinstance(value, str):
                value = re.sub(r'[R$£€,\s]', '', value)
            return float(value)
        except:
            return None

    elif data_type == 'int':
        try:
            if isinstance(value, str):
                value = re.sub(r'[,\s]', '', value)
            return int(float(value))
        except:
            return None

    else:  # string
        return str(value).strip()

def calculate_vat(excl_price: float, vat_rate: float = 0.15) -> float:
    """Calculate VAT-inclusive price"""
    if pd.isna(excl_price) or excl_price is None:
        return None
    return round(float(excl_price) * (1 + vat_rate), 2)

def calculate_excl(incl_price: float, vat_rate: float = 0.15) -> float:
    """Calculate VAT-exclusive price from inclusive"""
    if pd.isna(incl_price) or incl_price is None:
        return None
    return round(float(incl_price) / (1 + vat_rate), 2)

# ═══════════════════════════════════════════════════════════════════
# CORE PROCESSING FUNCTIONS
# ═══════════════════════════════════════════════════════════════════

def load_supplier_file(file_path: Path, config: Dict) -> Tuple[pd.DataFrame, Dict]:
    """Load and prepare supplier file"""
    stats = {
        'file': file_path.name,
        'supplier': config['supplier'],
        'raw_rows': 0,
        'header_row': 0,
        'data_rows': 0,
        'errors': []
    }

    try:
        # Try reading with different engines
        try:
            df = pd.read_excel(file_path, sheet_name=config['sheet'], header=None)
        except Exception as e:
            # Try openpyxl engine for large files
            df = pd.read_excel(file_path, sheet_name=config['sheet'], header=None, engine='openpyxl')

        stats['raw_rows'] = len(df)

        # Find header row
        header_row = find_header_row(df, config['mapping'])

        if header_row is None:
            # Use first row as header
            header_row = config.get('skip_rows', 0)

        stats['header_row'] = header_row

        # Re-read with correct header
        df = pd.read_excel(file_path, sheet_name=config['sheet'], header=header_row)

        # Remove completely empty rows
        df = df.dropna(how='all')

        stats['data_rows'] = len(df)

        return df, stats

    except Exception as e:
        stats['errors'].append(f"Load error: {str(e)}")
        return None, stats

def transform_to_master(df: pd.DataFrame, config: Dict, stats: Dict) -> pd.DataFrame:
    """Transform supplier data to master template format"""

    # Map columns
    column_map = map_columns(df, config['mapping'])

    if not column_map:
        stats['errors'].append("No column mappings found")
        return None

    stats['mapped_columns'] = len(column_map)

    # Create master dataframe
    master_df = pd.DataFrame(columns=MASTER_COLUMNS)

    # Map and sanitize data
    for source_col, master_col in column_map.items():
        if source_col in df.columns:
            if master_col in ['Cost Price Excl', 'Cost Price Incl', 'Retail Price Incl', 'Supplier Price Incl']:
                master_df[master_col] = df[source_col].apply(lambda x: sanitize_value(x, 'numeric'))
            elif master_col in ['QTY On Hand', 'Carton Quantity']:
                master_df[master_col] = df[source_col].apply(lambda x: sanitize_value(x, 'int'))
            else:
                master_df[master_col] = df[source_col].apply(lambda x: sanitize_value(x, 'str'))

    # Add supplier name
    master_df['Supplier'] = config['supplier']

    # Calculate missing VAT prices
    if 'Cost Price Excl' in master_df.columns and 'Cost Price Incl' not in master_df.columns:
        master_df['Cost Price Incl'] = master_df['Cost Price Excl'].apply(calculate_vat)
    elif 'Cost Price Incl' in master_df.columns and 'Cost Price Excl' not in master_df.columns:
        master_df['Cost Price Excl'] = master_df['Cost Price Incl'].apply(calculate_excl)

    # Add date captured
    master_df['Date Captured'] = datetime.now().strftime('%Y-%m-%d')

    # Remove rows with missing required fields
    initial_rows = len(master_df)
    for field in REQUIRED_FIELDS:
        master_df = master_df[master_df[field].notna()]

    stats['valid_rows'] = len(master_df)
    stats['rejected_rows'] = initial_rows - len(master_df)

    return master_df

def validate_data(df: pd.DataFrame, stats: Dict) -> List[str]:
    """Run validation checks on transformed data"""
    issues = []

    # Check required fields
    for field in REQUIRED_FIELDS:
        null_count = df[field].isna().sum()
        if null_count > 0:
            issues.append(f"⚠️ {null_count} rows missing {field}")

    # Check price validity
    if 'Cost Price Excl' in df.columns:
        negative = (df['Cost Price Excl'] < 0).sum()
        if negative > 0:
            issues.append(f"⚠️ {negative} rows with negative prices")

        zero = (df['Cost Price Excl'] == 0).sum()
        if zero > 0:
            issues.append(f"⚠️ {zero} rows with zero prices")

    # Check for duplicates
    if 'SKU' in df.columns:
        dupes = df['SKU'].duplicated().sum()
        if dupes > 0:
            issues.append(f"⚠️ {dupes} duplicate SKUs")

    # Data quality stats
    stats['null_percentages'] = {
        col: f"{(df[col].isna().sum() / len(df) * 100):.1f}%"
        for col in MASTER_COLUMNS if col in df.columns
    }

    return issues

def write_to_consolidated(master_df: pd.DataFrame, supplier_name: str,
                          consolidated_path: Path) -> Dict:
    """Write transformed data to consolidated workbook"""
    result = {
        'success': False,
        'rows_written': 0,
        'errors': []
    }

    try:
        # Load or create workbook
        try:
            wb = openpyxl.load_workbook(consolidated_path)
        except FileNotFoundError:
            wb = openpyxl.Workbook()
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

        # Create or clear supplier sheet
        if supplier_name in wb.sheetnames:
            wb.remove(wb[supplier_name])

        ws = wb.create_sheet(supplier_name)

        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(master_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        result['rows_written'] = len(master_df)

        # Update or create Master sheet
        if 'Master' not in wb.sheetnames:
            master_ws = wb.create_sheet('Master', 0)
            # Write headers
            for c_idx, col in enumerate(MASTER_COLUMNS, 1):
                master_ws.cell(row=1, column=c_idx, value=col)
        else:
            master_ws = wb['Master']

        # Append to master (after validation)
        start_row = master_ws.max_row + 1
        for r_idx, row in enumerate(dataframe_to_rows(master_df, index=False, header=False), start_row):
            for c_idx, value in enumerate(row, 1):
                master_ws.cell(row=r_idx, column=c_idx, value=value)

        wb.save(consolidated_path)
        result['success'] = True

    except Exception as e:
        result['errors'].append(f"Write error: {str(e)}")

    return result

# ═══════════════════════════════════════════════════════════════════
# MAIN PROCESSING PIPELINE
# ═══════════════════════════════════════════════════════════════════

def process_supplier(file_path: Path, config: Dict, consolidated_path: Path) -> Dict:
    """Complete processing pipeline for one supplier"""

    print(f"\n{'='*70}")
    print(f"PROCESSING: {config['supplier']}")
    print(f"File: {file_path.name}")
    print(f"{'='*70}")

    # Stage 1: Load
    print("\n[1/5] Loading file...")
    df, stats = load_supplier_file(file_path, config)

    if df is None:
        print(f"❌ Failed to load file: {stats.get('errors', [])}")
        return stats

    print(f"✅ Loaded {stats['data_rows']} rows (header at row {stats['header_row']})")

    # Stage 2: Transform
    print("\n[2/5] Transforming data...")
    master_df = transform_to_master(df, config, stats)

    if master_df is None or len(master_df) == 0:
        print(f"❌ Transformation failed: {stats.get('errors', [])}")
        return stats

    print(f"✅ Transformed to {stats['valid_rows']} valid rows ({stats.get('rejected_rows', 0)} rejected)")
    print(f"   Mapped {stats.get('mapped_columns', 0)} columns")

    # Stage 3: Validate
    print("\n[3/5] Validating data...")
    issues = validate_data(master_df, stats)

    if issues:
        print("⚠️ Validation warnings:")
        for issue in issues:
            print(f"   {issue}")
    else:
        print("✅ All validation checks passed")

    # Stage 4: Write to consolidated
    print("\n[4/5] Writing to consolidated workbook...")
    write_result = write_to_consolidated(master_df, config['supplier'], consolidated_path)

    if write_result['success']:
        print(f"✅ Wrote {write_result['rows_written']} rows to '{config['supplier']}' sheet")
        print(f"✅ Appended to 'Master' sheet")
    else:
        print(f"❌ Write failed: {write_result.get('errors', [])}")
        stats['errors'].extend(write_result.get('errors', []))

    stats.update(write_result)

    # Stage 5: Summary
    print("\n[5/5] Summary:")
    print(f"   Raw rows: {stats['raw_rows']}")
    print(f"   Valid rows: {stats['valid_rows']}")
    print(f"   Rejection rate: {(stats.get('rejected_rows', 0) / stats['raw_rows'] * 100):.1f}%")

    return stats

def main():
    """Process all files in Batch 2"""

    # Paths
    base_dir = Path('/mnt/k/00Project/MantisNXT')
    source_dir = base_dir / 'database/Uploads/drive-download-20250904T012253Z-1-001'
    consolidated_path = base_dir / 'database/Uploads/Consolidated_Supplier_Data.xlsx'

    print("=" * 70)
    print("SUPPLIER DATA PROCESSING - BATCH 2")
    print("=" * 70)
    print(f"Processing 7 suppliers: MD Distribution → Sennheiser")
    print(f"Source: {source_dir}")
    print(f"Output: {consolidated_path}")

    # Process each file
    all_stats = []

    for filename, config in BATCH_2_CONFIGS.items():
        file_path = source_dir / filename

        if not file_path.exists():
            print(f"\n❌ File not found: {filename}")
            continue

        try:
            stats = process_supplier(file_path, config, consolidated_path)
            all_stats.append(stats)
        except Exception as e:
            print(f"\n❌ CRITICAL ERROR processing {filename}: {str(e)}")
            import traceback
            traceback.print_exc()

    # Final report
    print("\n" + "=" * 70)
    print("BATCH 2 PROCESSING COMPLETE")
    print("=" * 70)

    total_raw = sum(s.get('raw_rows', 0) for s in all_stats)
    total_valid = sum(s.get('valid_rows', 0) for s in all_stats)
    total_rejected = sum(s.get('rejected_rows', 0) for s in all_stats)

    print(f"\nTotal Statistics:")
    print(f"   Files processed: {len(all_stats)}")
    print(f"   Raw rows: {total_raw:,}")
    print(f"   Valid rows: {total_valid:,}")
    print(f"   Rejected rows: {total_rejected:,}")
    print(f"   Success rate: {(total_valid / total_raw * 100):.1f}%")

    print(f"\nDetailed Results:")
    for stat in all_stats:
        status = "✅" if stat.get('success', False) else "❌"
        print(f"   {status} {stat['supplier']:.<30} {stat.get('valid_rows', 0):>6,} rows")

    print(f"\n✅ Output saved to: {consolidated_path}")

if __name__ == '__main__':
    main()
