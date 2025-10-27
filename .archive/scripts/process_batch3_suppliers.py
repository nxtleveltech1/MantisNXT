#!/usr/bin/env python3
"""
Process Batch 3 Supplier Files (Files 15-21)
Handles: Sonic Informed, Stage Audio Works, Stage One, Tuerk Multimedia,
         Tuerk Technologies, Viva Afrika, Yamaha
"""

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path
import re
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

# Paths
SOURCE_DIR = Path("/mnt/k/00Project/MantisNXT/database/Uploads/drive-download-20250904T012253Z-1-001")
CONSOLIDATED_FILE = Path("/mnt/k/00Project/MantisNXT/database/Uploads/Consolidated_Supplier_Data.xlsx")

# Master schema columns
MASTER_COLUMNS = [
    'Supplier', 'Brand', 'Model', 'Description', 'Category', 'Subcategory',
    'Cost Price (ZAR)', 'RRP (ZAR)', 'Stock Status', 'Lead Time (Days)',
    'Warranty (Months)', 'SKU', 'Barcode', 'Weight (kg)', 'Dimensions (cm)',
    'Supplier Code', 'Date Updated', 'Notes', 'Source File'
]

# Batch 3 files
BATCH_3_FILES = [
    {
        'file': 'SonicInformed_20250808_I.xlsx',
        'supplier': 'Sonic Informed',
        'expected_format': 'standard'
    },
    {
        'file': 'Stage Audio Works SOH info_20250826_0247.xlsx',
        'supplier': 'Stage Audio Works',
        'expected_format': 'stock_info'
    },
    {
        'file': 'Stage one FullPL-DP (7).xlsx',
        'supplier': 'Stage One Distribution',
        'expected_format': 'full_pl'
    },
    {
        'file': 'Tuerk Multimedia  Studio Price List_ JULY 2025.xlsx',
        'supplier': 'Tuerk Multimedia',
        'expected_format': 'standard'
    },
    {
        'file': 'Tuerk Tech General Price List JULY 2025.xls',
        'supplier': 'Tuerk Technologies',
        'expected_format': 'standard'
    },
    {
        'file': 'Viva Afrika Dealer Price List 07 May 2025.xlsx',
        'supplier': 'Viva Afrika',
        'expected_format': 'dealer_price'
    },
    {
        'file': 'YAMAHA RETAIL PRICELIST NOV 2024 (2).xlsx',
        'supplier': 'Yamaha',
        'expected_format': 'retail'
    }
]

def clean_currency_value(value):
    """Convert currency strings to float"""
    if pd.isna(value) or value == '':
        return None
    if isinstance(value, (int, float)):
        return float(value)
    # Remove currency symbols, spaces, commas
    cleaned = str(value).replace('R', '').replace('$', '').replace(',', '').replace(' ', '').strip()
    try:
        return float(cleaned) if cleaned else None
    except:
        return None

def clean_text(value):
    """Clean text fields"""
    if pd.isna(value) or value == '':
        return None
    return str(value).strip()

def find_header_row(df, header_keywords=['brand', 'model', 'description', 'price', 'sku', 'code', 'product']):
    """Find the row containing headers"""
    for idx, row in df.iterrows():
        row_text = ' '.join([str(cell).lower() for cell in row if pd.notna(cell)])
        if any(keyword in row_text for keyword in header_keywords):
            return idx
    return 0

def extract_brand_from_text(text):
    """Extract brand from product description or model"""
    if pd.isna(text):
        return None
    text = str(text).strip()
    # Common brand patterns
    brands = ['Yamaha', 'Roland', 'Korg', 'Behringer', 'Shure', 'Sennheiser',
              'Audio-Technica', 'AKG', 'Neumann', 'Focusrite', 'PreSonus',
              'Pioneer', 'Native Instruments', 'Arturia', 'Novation', 'Akai']

    for brand in brands:
        if brand.lower() in text.lower():
            return brand

    # Try to extract first word if it looks like a brand
    first_word = text.split()[0] if text.split() else None
    if first_word and len(first_word) > 2:
        return first_word

    return None

def process_sonic_informed(file_path):
    """Process Sonic Informed file"""
    print(f"\nProcessing: {file_path.name}")

    try:
        # Try reading different sheets
        xl_file = pd.ExcelFile(file_path)
        print(f"  Sheets found: {xl_file.sheet_names}")

        df = pd.read_excel(file_path, sheet_name=0)
        print(f"  Initial shape: {df.shape}")

        # Find header row
        header_row = find_header_row(df)
        if header_row > 0:
            df = pd.read_excel(file_path, sheet_name=0, header=header_row)
            print(f"  Found header at row {header_row}, new shape: {df.shape}")

        # Display column names to understand structure
        print(f"  Columns: {df.columns.tolist()[:10]}")

        # Map columns based on common patterns
        records = []
        for idx, row in df.iterrows():
            # Skip empty rows
            if row.isna().all():
                continue

            record = {col: None for col in MASTER_COLUMNS}
            record['Supplier'] = 'Sonic Informed'
            record['Source File'] = file_path.name
            record['Date Updated'] = datetime.now().strftime('%Y-%m-%d')

            # Try to map columns intelligently
            for col in df.columns:
                col_lower = str(col).lower()

                if 'brand' in col_lower or 'make' in col_lower:
                    record['Brand'] = clean_text(row[col])
                elif 'model' in col_lower or 'part' in col_lower:
                    record['Model'] = clean_text(row[col])
                elif 'description' in col_lower or 'product' in col_lower:
                    record['Description'] = clean_text(row[col])
                elif 'category' in col_lower or 'type' in col_lower:
                    record['Category'] = clean_text(row[col])
                elif 'cost' in col_lower or 'dealer' in col_lower:
                    record['Cost Price (ZAR)'] = clean_currency_value(row[col])
                elif 'rrp' in col_lower or 'retail' in col_lower:
                    record['RRP (ZAR)'] = clean_currency_value(row[col])
                elif 'sku' in col_lower or 'code' in col_lower:
                    record['SKU'] = clean_text(row[col])
                elif 'stock' in col_lower:
                    record['Stock Status'] = clean_text(row[col])

            # Extract brand if not found
            if not record['Brand'] and record['Description']:
                record['Brand'] = extract_brand_from_text(record['Description'])

            records.append(record)

        result_df = pd.DataFrame(records)
        print(f"  Extracted {len(result_df)} records")
        return result_df

    except Exception as e:
        print(f"  ERROR: {str(e)}")
        return pd.DataFrame(columns=MASTER_COLUMNS)

def process_stage_audio_works(file_path):
    """Process Stage Audio Works SOH file"""
    print(f"\nProcessing: {file_path.name}")

    try:
        xl_file = pd.ExcelFile(file_path)
        print(f"  Sheets found: {xl_file.sheet_names}")

        df = pd.read_excel(file_path, sheet_name=0)
        print(f"  Initial shape: {df.shape}")

        header_row = find_header_row(df)
        if header_row > 0:
            df = pd.read_excel(file_path, sheet_name=0, header=header_row)

        print(f"  Columns: {df.columns.tolist()[:10]}")

        records = []
        for idx, row in df.iterrows():
            if row.isna().all():
                continue

            record = {col: None for col in MASTER_COLUMNS}
            record['Supplier'] = 'Stage Audio Works'
            record['Source File'] = file_path.name
            record['Date Updated'] = datetime.now().strftime('%Y-%m-%d')

            for col in df.columns:
                col_lower = str(col).lower()

                if 'brand' in col_lower:
                    record['Brand'] = clean_text(row[col])
                elif 'model' in col_lower or 'part' in col_lower:
                    record['Model'] = clean_text(row[col])
                elif 'description' in col_lower:
                    record['Description'] = clean_text(row[col])
                elif 'category' in col_lower:
                    record['Category'] = clean_text(row[col])
                elif 'price' in col_lower or 'cost' in col_lower:
                    record['Cost Price (ZAR)'] = clean_currency_value(row[col])
                elif 'sku' in col_lower or 'code' in col_lower:
                    record['SKU'] = clean_text(row[col])
                elif 'stock' in col_lower or 'qty' in col_lower or 'soh' in col_lower:
                    stock_val = clean_text(row[col])
                    if stock_val:
                        try:
                            qty = int(float(stock_val))
                            record['Stock Status'] = 'In Stock' if qty > 0 else 'Out of Stock'
                        except:
                            record['Stock Status'] = stock_val

            if not record['Brand'] and record['Description']:
                record['Brand'] = extract_brand_from_text(record['Description'])

            records.append(record)

        result_df = pd.DataFrame(records)
        print(f"  Extracted {len(result_df)} records")
        return result_df

    except Exception as e:
        print(f"  ERROR: {str(e)}")
        return pd.DataFrame(columns=MASTER_COLUMNS)

def process_stage_one(file_path):
    """Process Stage One Distribution full price list"""
    print(f"\nProcessing: {file_path.name}")

    try:
        xl_file = pd.ExcelFile(file_path)
        print(f"  Sheets found: {xl_file.sheet_names}")

        df = pd.read_excel(file_path, sheet_name=0)
        print(f"  Initial shape: {df.shape}")

        header_row = find_header_row(df)
        if header_row > 0:
            df = pd.read_excel(file_path, sheet_name=0, header=header_row)

        print(f"  Columns: {df.columns.tolist()[:10]}")

        records = []
        for idx, row in df.iterrows():
            if row.isna().all():
                continue

            record = {col: None for col in MASTER_COLUMNS}
            record['Supplier'] = 'Stage One Distribution'
            record['Source File'] = file_path.name
            record['Date Updated'] = datetime.now().strftime('%Y-%m-%d')

            for col in df.columns:
                col_lower = str(col).lower()

                if 'brand' in col_lower or 'make' in col_lower:
                    record['Brand'] = clean_text(row[col])
                elif 'model' in col_lower or 'part' in col_lower:
                    record['Model'] = clean_text(row[col])
                elif 'description' in col_lower:
                    record['Description'] = clean_text(row[col])
                elif 'category' in col_lower:
                    record['Category'] = clean_text(row[col])
                elif 'dealer' in col_lower or 'dp' in col_lower:
                    record['Cost Price (ZAR)'] = clean_currency_value(row[col])
                elif 'rrp' in col_lower or 'retail' in col_lower:
                    record['RRP (ZAR)'] = clean_currency_value(row[col])
                elif 'sku' in col_lower or 'code' in col_lower:
                    record['SKU'] = clean_text(row[col])
                elif 'stock' in col_lower:
                    record['Stock Status'] = clean_text(row[col])

            if not record['Brand'] and record['Description']:
                record['Brand'] = extract_brand_from_text(record['Description'])

            records.append(record)

        result_df = pd.DataFrame(records)
        print(f"  Extracted {len(result_df)} records")
        return result_df

    except Exception as e:
        print(f"  ERROR: {str(e)}")
        return pd.DataFrame(columns=MASTER_COLUMNS)

def process_tuerk_multimedia(file_path):
    """Process Tuerk Multimedia Studio Price List"""
    print(f"\nProcessing: {file_path.name}")

    try:
        xl_file = pd.ExcelFile(file_path)
        print(f"  Sheets found: {xl_file.sheet_names}")

        # Try to find the main price list sheet
        target_sheet = xl_file.sheet_names[0]
        for sheet_name in xl_file.sheet_names:
            if any(keyword in sheet_name.lower() for keyword in ['price', 'list', 'products']):
                target_sheet = sheet_name
                break

        df = pd.read_excel(file_path, sheet_name=target_sheet)
        print(f"  Using sheet: {target_sheet}, shape: {df.shape}")

        header_row = find_header_row(df)
        if header_row > 0:
            df = pd.read_excel(file_path, sheet_name=target_sheet, header=header_row)

        print(f"  Columns: {df.columns.tolist()[:10]}")

        records = []
        for idx, row in df.iterrows():
            if row.isna().all():
                continue

            record = {col: None for col in MASTER_COLUMNS}
            record['Supplier'] = 'Tuerk Multimedia'
            record['Source File'] = file_path.name
            record['Date Updated'] = datetime.now().strftime('%Y-%m-%d')

            for col in df.columns:
                col_lower = str(col).lower()

                if 'brand' in col_lower:
                    record['Brand'] = clean_text(row[col])
                elif 'model' in col_lower or 'part' in col_lower:
                    record['Model'] = clean_text(row[col])
                elif 'description' in col_lower:
                    record['Description'] = clean_text(row[col])
                elif 'category' in col_lower:
                    record['Category'] = clean_text(row[col])
                elif 'cost' in col_lower or 'dealer' in col_lower:
                    record['Cost Price (ZAR)'] = clean_currency_value(row[col])
                elif 'rrp' in col_lower or 'retail' in col_lower:
                    record['RRP (ZAR)'] = clean_currency_value(row[col])
                elif 'sku' in col_lower or 'code' in col_lower:
                    record['SKU'] = clean_text(row[col])
                elif 'stock' in col_lower:
                    record['Stock Status'] = clean_text(row[col])

            if not record['Brand'] and record['Description']:
                record['Brand'] = extract_brand_from_text(record['Description'])

            records.append(record)

        result_df = pd.DataFrame(records)
        print(f"  Extracted {len(result_df)} records")
        return result_df

    except Exception as e:
        print(f"  ERROR: {str(e)}")
        return pd.DataFrame(columns=MASTER_COLUMNS)

def process_tuerk_technologies(file_path):
    """Process Tuerk Technologies General Price List (.xls)"""
    print(f"\nProcessing: {file_path.name}")

    try:
        # Handle .xls format
        xl_file = pd.ExcelFile(file_path, engine='xlrd')
        print(f"  Sheets found: {xl_file.sheet_names}")

        target_sheet = xl_file.sheet_names[0]
        df = pd.read_excel(file_path, sheet_name=target_sheet, engine='xlrd')
        print(f"  Using sheet: {target_sheet}, shape: {df.shape}")

        header_row = find_header_row(df)
        if header_row > 0:
            df = pd.read_excel(file_path, sheet_name=target_sheet, header=header_row, engine='xlrd')

        print(f"  Columns: {df.columns.tolist()[:10]}")

        records = []
        for idx, row in df.iterrows():
            if row.isna().all():
                continue

            record = {col: None for col in MASTER_COLUMNS}
            record['Supplier'] = 'Tuerk Technologies'
            record['Source File'] = file_path.name
            record['Date Updated'] = datetime.now().strftime('%Y-%m-%d')

            for col in df.columns:
                col_lower = str(col).lower()

                if 'brand' in col_lower:
                    record['Brand'] = clean_text(row[col])
                elif 'model' in col_lower or 'part' in col_lower:
                    record['Model'] = clean_text(row[col])
                elif 'description' in col_lower:
                    record['Description'] = clean_text(row[col])
                elif 'category' in col_lower:
                    record['Category'] = clean_text(row[col])
                elif 'cost' in col_lower or 'dealer' in col_lower:
                    record['Cost Price (ZAR)'] = clean_currency_value(row[col])
                elif 'rrp' in col_lower or 'retail' in col_lower:
                    record['RRP (ZAR)'] = clean_currency_value(row[col])
                elif 'sku' in col_lower or 'code' in col_lower:
                    record['SKU'] = clean_text(row[col])
                elif 'stock' in col_lower:
                    record['Stock Status'] = clean_text(row[col])

            if not record['Brand'] and record['Description']:
                record['Brand'] = extract_brand_from_text(record['Description'])

            records.append(record)

        result_df = pd.DataFrame(records)
        print(f"  Extracted {len(result_df)} records")
        return result_df

    except Exception as e:
        print(f"  ERROR: {str(e)}")
        return pd.DataFrame(columns=MASTER_COLUMNS)

def process_viva_afrika(file_path):
    """Process Viva Afrika Dealer Price List"""
    print(f"\nProcessing: {file_path.name}")

    try:
        xl_file = pd.ExcelFile(file_path)
        print(f"  Sheets found: {xl_file.sheet_names}")

        # Process all relevant sheets
        all_records = []

        for sheet_name in xl_file.sheet_names:
            if 'info' in sheet_name.lower() or 'index' in sheet_name.lower():
                continue

            print(f"  Processing sheet: {sheet_name}")
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            print(f"    Shape: {df.shape}")

            header_row = find_header_row(df)
            if header_row > 0:
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)

            for idx, row in df.iterrows():
                if row.isna().all():
                    continue

                record = {col: None for col in MASTER_COLUMNS}
                record['Supplier'] = 'Viva Afrika'
                record['Source File'] = file_path.name
                record['Date Updated'] = datetime.now().strftime('%Y-%m-%d')

                # Use sheet name as category if no explicit category
                if sheet_name not in ['Sheet1', 'Sheet2']:
                    record['Category'] = sheet_name

                for col in df.columns:
                    col_lower = str(col).lower()

                    if 'brand' in col_lower:
                        record['Brand'] = clean_text(row[col])
                    elif 'model' in col_lower or 'part' in col_lower:
                        record['Model'] = clean_text(row[col])
                    elif 'description' in col_lower:
                        record['Description'] = clean_text(row[col])
                    elif 'category' in col_lower:
                        record['Category'] = clean_text(row[col])
                    elif 'dealer' in col_lower or 'cost' in col_lower:
                        record['Cost Price (ZAR)'] = clean_currency_value(row[col])
                    elif 'rrp' in col_lower or 'retail' in col_lower:
                        record['RRP (ZAR)'] = clean_currency_value(row[col])
                    elif 'sku' in col_lower or 'code' in col_lower:
                        record['SKU'] = clean_text(row[col])
                    elif 'stock' in col_lower:
                        record['Stock Status'] = clean_text(row[col])

                if not record['Brand'] and record['Description']:
                    record['Brand'] = extract_brand_from_text(record['Description'])

                all_records.append(record)

        result_df = pd.DataFrame(all_records)
        print(f"  Total extracted: {len(result_df)} records")
        return result_df

    except Exception as e:
        print(f"  ERROR: {str(e)}")
        return pd.DataFrame(columns=MASTER_COLUMNS)

def process_yamaha(file_path):
    """Process Yamaha Retail Pricelist"""
    print(f"\nProcessing: {file_path.name}")

    try:
        xl_file = pd.ExcelFile(file_path)
        print(f"  Sheets found: {xl_file.sheet_names}")

        all_records = []

        for sheet_name in xl_file.sheet_names:
            if 'info' in sheet_name.lower() or 'index' in sheet_name.lower():
                continue

            print(f"  Processing sheet: {sheet_name}")
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            print(f"    Shape: {df.shape}")

            header_row = find_header_row(df)
            if header_row > 0:
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)

            for idx, row in df.iterrows():
                if row.isna().all():
                    continue

                record = {col: None for col in MASTER_COLUMNS}
                record['Supplier'] = 'Yamaha'
                record['Brand'] = 'Yamaha'  # All products are Yamaha
                record['Source File'] = file_path.name
                record['Date Updated'] = datetime.now().strftime('%Y-%m-%d')

                # Use sheet name as category
                if sheet_name not in ['Sheet1', 'Sheet2']:
                    record['Category'] = sheet_name

                for col in df.columns:
                    col_lower = str(col).lower()

                    if 'model' in col_lower or 'part' in col_lower:
                        record['Model'] = clean_text(row[col])
                    elif 'description' in col_lower or 'product' in col_lower:
                        record['Description'] = clean_text(row[col])
                    elif 'category' in col_lower:
                        record['Category'] = clean_text(row[col])
                    elif 'cost' in col_lower or 'dealer' in col_lower:
                        record['Cost Price (ZAR)'] = clean_currency_value(row[col])
                    elif 'rrp' in col_lower or 'retail' in col_lower:
                        record['RRP (ZAR)'] = clean_currency_value(row[col])
                    elif 'sku' in col_lower or 'code' in col_lower:
                        record['SKU'] = clean_text(row[col])
                    elif 'stock' in col_lower:
                        record['Stock Status'] = clean_text(row[col])

                all_records.append(record)

        result_df = pd.DataFrame(all_records)
        print(f"  Total extracted: {len(result_df)} records")
        return result_df

    except Exception as e:
        print(f"  ERROR: {str(e)}")
        return pd.DataFrame(columns=MASTER_COLUMNS)

def validate_data(df, supplier_name):
    """Validate processed data quality"""
    print(f"\n=== Validation: {supplier_name} ===")

    stats = {
        'total_records': len(df),
        'has_brand': df['Brand'].notna().sum(),
        'has_model': df['Model'].notna().sum(),
        'has_description': df['Description'].notna().sum(),
        'has_cost_price': df['Cost Price (ZAR)'].notna().sum(),
        'has_rrp': df['RRP (ZAR)'].notna().sum(),
        'has_sku': df['SKU'].notna().sum(),
        'completeness': 0
    }

    # Calculate completeness score
    if stats['total_records'] > 0:
        key_fields = ['Brand', 'Description', 'Cost Price (ZAR)']
        completeness_scores = [df[field].notna().sum() / stats['total_records'] * 100 for field in key_fields]
        stats['completeness'] = sum(completeness_scores) / len(completeness_scores)

    print(f"  Total Records: {stats['total_records']}")
    print(f"  Has Brand: {stats['has_brand']} ({stats['has_brand']/stats['total_records']*100:.1f}%)")
    print(f"  Has Model: {stats['has_model']} ({stats['has_model']/stats['total_records']*100:.1f}%)")
    print(f"  Has Description: {stats['has_description']} ({stats['has_description']/stats['total_records']*100:.1f}%)")
    print(f"  Has Cost Price: {stats['has_cost_price']} ({stats['has_cost_price']/stats['total_records']*100:.1f}%)")
    print(f"  Has RRP: {stats['has_rrp']} ({stats['has_rrp']/stats['total_records']*100:.1f}%)")
    print(f"  Has SKU: {stats['has_sku']} ({stats['has_sku']/stats['total_records']*100:.1f}%)")
    print(f"  Completeness Score: {stats['completeness']:.1f}%")

    return stats

def append_to_consolidated(df, supplier_name):
    """Append processed data to Consolidated workbook"""
    print(f"\n=== Appending to Consolidated: {supplier_name} ===")

    try:
        # Load existing workbook
        wb = openpyxl.load_workbook(CONSOLIDATED_FILE)

        # Create supplier-specific sheet
        sheet_name = supplier_name[:31]  # Excel sheet name limit
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)

        # Write headers
        for col_idx, col_name in enumerate(MASTER_COLUMNS, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # Write data
        for row_idx, record in enumerate(df.to_dict('records'), start=2):
            for col_idx, col_name in enumerate(MASTER_COLUMNS, start=1):
                ws.cell(row=row_idx, column=col_idx, value=record.get(col_name))

        print(f"  Created sheet '{sheet_name}' with {len(df)} rows")

        # Append to Master sheet
        if 'Master' in wb.sheetnames:
            master_ws = wb['Master']
            start_row = master_ws.max_row + 1

            for row_idx, record in enumerate(df.to_dict('records'), start=start_row):
                for col_idx, col_name in enumerate(MASTER_COLUMNS, start=1):
                    master_ws.cell(row=row_idx, column=col_idx, value=record.get(col_name))

            print(f"  Appended {len(df)} rows to Master sheet (starting at row {start_row})")
        else:
            print("  WARNING: Master sheet not found")

        # Save workbook
        wb.save(CONSOLIDATED_FILE)
        print(f"  Saved to: {CONSOLIDATED_FILE}")

        return True

    except Exception as e:
        print(f"  ERROR appending to consolidated: {str(e)}")
        return False

def main():
    """Main processing function"""
    print("="*80)
    print("BATCH 3 SUPPLIER FILE PROCESSING")
    print("Files 15-21: Sonic Informed → Yamaha")
    print("="*80)

    processors = {
        'Sonic Informed': process_sonic_informed,
        'Stage Audio Works': process_stage_audio_works,
        'Stage One Distribution': process_stage_one,
        'Tuerk Multimedia': process_tuerk_multimedia,
        'Tuerk Technologies': process_tuerk_technologies,
        'Viva Afrika': process_viva_afrika,
        'Yamaha': process_yamaha
    }

    overall_stats = []

    for file_info in BATCH_3_FILES:
        file_path = SOURCE_DIR / file_info['file']
        supplier = file_info['supplier']

        if not file_path.exists():
            print(f"\nWARNING: File not found: {file_path}")
            continue

        print(f"\n{'='*80}")
        print(f"Processing: {supplier}")
        print(f"File: {file_path.name}")
        print(f"{'='*80}")

        # Process file
        processor = processors.get(supplier)
        if processor:
            df = processor(file_path)

            if not df.empty:
                # Validate
                stats = validate_data(df, supplier)
                stats['supplier'] = supplier
                stats['file'] = file_info['file']
                overall_stats.append(stats)

                # Append to consolidated
                success = append_to_consolidated(df, supplier)

                if success:
                    print(f"\n✅ SUCCESS: {supplier} processed and appended")
                else:
                    print(f"\n❌ FAILED: Could not append {supplier}")
            else:
                print(f"\n❌ FAILED: No data extracted from {supplier}")
        else:
            print(f"\n❌ ERROR: No processor defined for {supplier}")

    # Final summary
    print("\n" + "="*80)
    print("BATCH 3 PROCESSING SUMMARY")
    print("="*80)

    summary_df = pd.DataFrame(overall_stats)
    print("\nSupplier Statistics:")
    print(summary_df.to_string(index=False))

    total_records = summary_df['total_records'].sum()
    avg_completeness = summary_df['completeness'].mean()

    print(f"\n{'='*80}")
    print(f"TOTAL RECORDS PROCESSED: {total_records}")
    print(f"AVERAGE COMPLETENESS: {avg_completeness:.1f}%")
    print(f"{'='*80}")

if __name__ == '__main__':
    main()
